"""Testes dos relatórios (consultas + geração de arquivos)."""
import json
import re
import sqlite3
import sys
from datetime import date, timedelta
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
import licitarium
import relatorios


def _ler_planilha(caminho):
    """Linhas da planilha exportada, cabeçalho incluso — cada uma como
    tupla de valores das células, na ordem das colunas."""
    ws = openpyxl.load_workbook(caminho, read_only=True).active
    return [tuple(c.value for c in linha) for linha in ws.iter_rows()]


@pytest.fixture
def db():
    con = sqlite3.connect(":memory:")
    con.row_factory = sqlite3.Row
    con.executescript(licitarium.SCHEMA)
    raw_c = json.dumps({"amparoLegal": {"nome": "Art. 75, II"}})
    con.executemany(
        "INSERT INTO contratacoes (numero_controle, ano, sequencial,"
        " modalidade_nome, objeto, valor_estimado, valor_homologado,"
        " data_publicacao, raw) VALUES (?,?,?,?,?,?,?,?,?)",
        [("A", 2026, 1, "Dispensa", "Merenda", 100.0, 80.0,
          "2026-03-01", raw_c),
         ("B", 2026, 2, "Pregão", "Obras", 200.0, None, "2026-04-01", raw_c),
         ("C", 2025, 1, "Pregão", "Antigo", 50.0, 50.0, "2025-01-01", raw_c)])
    con.executemany(
        "INSERT INTO contratos (numero_controle, fornecedor_nome, objeto,"
        " valor_global, vigencia_inicio, vigencia_fim, data_publicacao, raw)"
        " VALUES (?,?,?,?,?,?,?,?)",
        [("CT1", "Fornecedor X", "Serviço", 1000.0, "2026-01-01", "2099-01-01",
          "2026-01-05", json.dumps({"numeroContratoEmpenho": "7/2026"})),
         ("CT2", "Fornecedor Y", "Vencido", 500.0, "2020-01-01", "2020-12-31",
          "2020-01-05", "{}")])
    con.execute(
        "INSERT INTO atas (numero_controle, contratacao_controle,"
        " vigencia_inicio, vigencia_fim, raw) VALUES (?,?,?,?,?)",
        ("AT1", "A", "2026-01-01", "2099-01-01",
         json.dumps({"numeroAtaRegistroPreco": "9", "anoAta": 2026,
                     "objetoContratacao": "RP merenda"})))
    con.commit()
    yield con
    con.close()


def test_data_now_do_sql_sempre_em_localtime():
    """date('now')/datetime('now') do SQLite são UTC. Entre ~21h e meia-noite
    de Brasília, o UTC já é o dia seguinte — um contrato vencendo hoje lia
    como já vencido nos painéis de vigência/prazo (auditoria 2026-08-11).
    Comportamento depende do fuso da máquina que roda o teste (não dá pra
    simular sem mockar as funções de data do próprio SQLite), então a
    garantia aqui é estática: nenhum 'now' bruto sobra no arquivo."""
    fonte = Path(relatorios.__file__).read_text(encoding="utf-8")
    nus = re.findall(r"(?:date|datetime)\('now'(?!,'localtime')[^)]*\)", fonte)
    assert not nus, f"'now' sem localtime: {nus}"


def test_dados_contratacoes_amparo_e_desagio(db):
    d = relatorios.dados_contratacoes(db, ano=2026)
    assert d["totais"]["n"] == 2
    assert d["linhas"][0]["amparo"] == "Art. 75, II"
    # deságio só sobre o processo com ambos os valores: 1 - 80/100 = 20%
    assert round(d["totais"]["desagio"], 1) == 20.0
    assert d["totais"]["estimado"] == 300.0


def test_dados_contratos_vigentes(db):
    assert relatorios.dados_contratos(db, vigentes=True)["totais"]["n"] == 1
    assert relatorios.dados_contratos(db, ano=2020)["totais"]["n"] == 1
    assert relatorios.dados_contratos(db)["totais"]["n"] == 2


def test_dados_executivo(db):
    d = relatorios.dados_executivo(db, 2026)
    assert d["cards"]["n"] == 2
    assert d["cards"]["contratos_vigentes"] == 1
    # contrato e ata com fim em 2099 não entram nos 90 dias
    assert d["vencendo"] == []
    assert d["meses"]["03"]["n"] == 1


def test_gerar_html_e_xlsx(db, tmp_path):
    r = relatorios.gerar(db, "contratacoes", {"ano": 2026},
                         "Testópolis", "SP", tmp_path)
    html = Path(r["html"]).read_text(encoding="utf-8")
    assert "Testópolis" in html and "MERENDA" not in html  # caixa alta é CSS
    assert "Merenda" in html and "Art. 75, II" in html
    assert "2 contratações" in html
    linhas = _ler_planilha(r["xlsx"])
    assert linhas[0][0] == "SEQUENCIAL"          # cabeçalho traduzido, caixa alta
    assert len(linhas) == 3                      # cabeçalho + 2 linhas


def test_escrever_planilha_traduz_cabecalho_e_formata_numero(tmp_path):
    """Motor central da exportação (pedido do usuário 2026-08-29: trocar o
    CSV cru por planilha "bonitinha"). Cabeçalho em CAIXA ALTA e negrito
    (2026-08-30, a partir de um modelo que o próprio usuário poliu à mão),
    número com formato de milhar — não texto cru."""
    import openpyxl as oxl
    caminho = tmp_path / "teste.xlsx"
    relatorios.escrever_planilha(
        caminho, [{"valor_homologado": 1234.5, "objeto_desconhecido_xyz": "Merenda"}])
    wb = oxl.load_workbook(caminho)
    ws = wb.active
    cab1, cab2 = ws["A1"], ws["B1"]
    assert cab1.value == "VALOR HOMOLOGADO"        # rótulo do dicionário, maiúsculo
    assert cab2.value == "OBJETO DESCONHECIDO XYZ"  # sem entrada: Title Case + maiúsculo
    assert cab1.font.bold is True
    assert ws["A2"].value == 1234.5                 # número real, não string
    assert ws["A2"].number_format == "#,##0.00"
    assert ws.freeze_panes == "A2"


def test_escrever_planilha_calcula_desagio_por_formula(tmp_path):
    """Modelo do usuário (2026-08-30): quando há par estimado/homologado, a
    planilha ganha coluna Deságio calculada por FÓRMULA do Excel — não um
    número congelado, senão editar estimado/homologado na planilha deixa o
    deságio mentindo."""
    import openpyxl as oxl
    caminho = tmp_path / "desagio.xlsx"
    relatorios.escrever_planilha(caminho, [
        {"objeto": "Merenda", "valor_estimado": 100.0, "valor_homologado": 80.0},
        {"objeto": "Sem homologação ainda", "valor_estimado": 50.0,
         "valor_homologado": None},
    ])
    wb = oxl.load_workbook(caminho)  # com fórmula, não data_only
    ws = wb.active
    cabecalho = [c.value for c in ws[1]]
    # Deságio entra logo depois do par (mesmo lugar do modelo do usuário:
    # colunas I/J estimado/homologado, K deságio — nunca entre os dois)
    assert cabecalho == ["OBJETO", "VALOR ESTIMADO", "VALOR HOMOLOGADO", "DESÁGIO"]
    assert ws["D2"].value == "=(B2-C2)/B2"
    assert ws["D2"].number_format == "0.00%"
    assert ws["D3"].value is None  # sem homologado, sem fórmula (não divide por nada)


def test_escrever_planilha_data_iso_vira_datetime_de_verdade(tmp_path):
    """Achado do usuário (2026-08-30, screenshot): publicação/encerramento
    saíam como texto ISO cru ("2026-08-26T11:19:47") — não dá pra Excel
    ordenar/filtrar como data assim. Campos data_*/vigencia_* viram
    datetime real, formatado; hora meia-noite mostra só a data."""
    import datetime as dt

    import openpyxl as oxl
    caminho = tmp_path / "datas.xlsx"
    relatorios.escrever_planilha(caminho, [
        {"objeto": "Merenda", "data_publicacao": "2026-08-26T11:19:47",
         "vigencia_fim": "2026-08-20T00:00:00", "numero_controle": "abc-1"},
    ])
    wb = oxl.load_workbook(caminho)
    ws = wb.active
    assert ws["B2"].value == dt.datetime(2026, 8, 26, 11, 19, 47)
    assert ws["B2"].number_format == "DD/MM/YYYY HH:MM"
    assert ws["C2"].value == dt.datetime(2026, 8, 20, 0, 0, 0)
    assert ws["C2"].number_format == "DD/MM/YYYY"       # meia-noite: só data
    assert ws["D2"].value == "=C2-HOJE()"                # vencimento (dias), auto
    assert ws["D2"].number_format == "0"
    assert ws["E2"].value == "abc-1"                     # não-data intocada


def test_gerar_executivo_sem_planilha(db, tmp_path):
    r = relatorios.gerar(db, "executivo", {"ano": 2026}, "T", "SP", tmp_path)
    assert r["xlsx"] is None
    assert "Resumo Executivo" in Path(r["html"]).read_text(encoding="utf-8")


def test_executivo_usa_grafico_pronto_quando_vem_da_tela(db, tmp_path):
    """`graficos` (ECharts que a tela já desenhou, achado 2026-08-11, mesmo
    padrão da pesquisa de preços) substitui `_grafico_meses`/`_grafico_barras`
    slot a slot; sem ele, o fallback hand-SVG segue funcionando."""
    marcador_meses = '<svg id="meses-de-mentirinha"></svg>'
    marcador_mod = '<svg id="modalidade-de-mentirinha"></svg>'
    r = relatorios.gerar(db, "executivo",
                         {"ano": 2026, "graficos": {"meses": marcador_meses,
                                                    "modalidade": marcador_mod}},
                         "T", "SP", tmp_path)
    html = Path(r["html"]).read_text(encoding="utf-8")
    assert marcador_meses in html and marcador_mod in html

    r2 = relatorios.gerar(db, "executivo", {"ano": 2026}, "T", "SP", tmp_path)
    html2 = Path(r2["html"]).read_text(encoding="utf-8")
    assert marcador_meses not in html2
    assert "<svg" in html2   # o fallback ainda desenha algo


def test_render_detalhe_usa_o_html_que_a_tela_ja_montou():
    """Ficha do modal de detalhe (achado 2026-08-12): `meta_html` vem
    pronto — rótulo/valor já formatados — e entra sem reimplementar nada
    em Python; só título e subtítulo passam pelo escape de sempre."""
    meta = ('<div><div class="k">Órgão</div>'
            '<div class="v">Prefeitura &amp; Câmara</div></div>')
    html = relatorios.render_detalhe(
        "Aquisição de <material>", "12345/2026-1", meta, "T", "SP")
    assert meta in html                        # HTML pronto, sem retoque
    assert "&lt;material&gt;" in html          # título passa pelo escape
    assert "12345/2026-1" in html
    assert 'size: A4 portrait' in html         # ficha de um registro é retrato


def test_render_detalhe_leva_o_json_colorido_do_modal():
    """achado 2026-08-12: a ficha impressa saía sem o "Dados completos"
    que o modal mostra — pedido do usuário, mesmo padrão de meta_html:
    a tela já colore (jsonColorido), o Python só encaixa."""
    raw = '<span class="j-chave">"objeto"</span>: <span class="j-str">"x"</span>'
    html = relatorios.render_detalhe("T", "1/2026", "<div></div>", "T", "SP",
                                     raw_html=raw)
    assert raw in html
    assert "Dados completos" in html

    sem_raw = relatorios.render_detalhe("T", "1/2026", "<div></div>", "T", "SP")
    assert "Dados completos" not in sem_raw    # sem raw_html, nem a seção aparece


def test_render_detalhe_cabecalho_e_municipio_objeto_desce_pro_corpo():
    """Pedido do usuário (2026-08-12): o cabeçalho é brasão + identificação
    do município (não o objeto, que é comprido demais pra ficar ali) —
    o objeto desce pro corpo, em caixa alta e justificado."""
    html = relatorios.render_detalhe(
        "objeto bem comprido de uma contratação qualquer", "1/2026",
        "<div></div>", "Orindiúva", "SP")
    # <title> e o h1 do cabeçalho mostram o município, não o objeto
    assert "<title>Orindiúva — SP</title>" in html
    assert "<h1>Orindiúva — SP</h1>" in html
    # o objeto vira parágrafo próprio, fora do cabeçalho, com a classe
    # que aplica caixa alta + justificado
    assert ('<p class="ficha-objeto">objeto bem comprido de uma '
            'contratação qualquer</p>') in html


def test_gerar_economia_sem_itens_nao_gera_planilha(db, tmp_path):
    """Sem item com par estimado/homologado, por_familia fica vazia — mesmo
    critério de "sem dados" que os outros relatórios já usam (linhas_csv
    vazia não gera arquivo)."""
    r = relatorios.gerar(db, "economia", {"ano": 2026}, "T", "SP", tmp_path)
    assert r["xlsx"] is None
    html = Path(r["html"]).read_text(encoding="utf-8")
    assert "Economia e Comparativos" in html
    # 300 estimados (A+B) - 80 homologados (só A) = 220
    assert "220,00" in html


def test_economia_usa_graficos_prontos_nos_quatro_slots(db, tmp_path):
    """Os 4 gráficos de barra (modalidade/família/categoria/fornecedor)
    aceitam o SVG pronto vindo da tela, cada um no seu slot."""
    marcadores = {k: f'<svg id="{k}-de-mentirinha"></svg>'
                 for k in ("modalidade", "familia", "categoria", "fornecedor")}
    r = relatorios.gerar(db, "economia", {"ano": 2026, "graficos": marcadores},
                         "T", "SP", tmp_path)
    html = Path(r["html"]).read_text(encoding="utf-8")
    for marcador in marcadores.values():
        assert marcador in html


def test_gerar_economia_com_itens_traz_familia_no_documento_e_na_planilha(
        db, tmp_path):
    db.execute(
        "INSERT INTO itens (id, contratacao_controle, ano, descricao,"
        " categoria, valor_total_estimado, valor_total_homologado,"
        " referencia, raw) VALUES ('A#1','A',2026,'MERENDA ESCOLAR',"
        " 'Alimentação',100.0,80.0,0,'{}')")
    db.commit()
    r = relatorios.gerar(db, "economia", {"ano": 2026}, "T", "SP", tmp_path)
    assert r["xlsx"] is not None
    linhas = _ler_planilha(r["xlsx"])
    assert any("MERENDA ESCOLAR" in str(v) for linha in linhas for v in linha)
    html = Path(r["html"]).read_text(encoding="utf-8")
    assert "MERENDA ESCOLAR" in html
    assert "Alimentação" in html            # tabela por categoria


def test_economia_por_modalidade_nao_corta_o_grafico(db, tmp_path):
    """O gráfico "Por modalidade" cortava em 8 (relatorios.py, achado da
    auditoria 2026-08-11) enquanto a tabela logo abaixo, na mesma página,
    e a vista Economia na tela mostram a lista inteira — um município com
    mais de 8 modalidades no exercício via gráfico e tabela discordando em
    quantidade de itens. Sem par estimado/homologado por linha independente
    e valor_estimado decrescente, a nona modalidade era exatamente a que
    sumia do gráfico."""
    db.executemany(
        "INSERT INTO contratacoes (numero_controle, ano, sequencial,"
        " modalidade_nome, objeto, valor_estimado, valor_homologado,"
        " data_publicacao, referencia) VALUES (?,2026,?,?,?,?,?,'2026-01-01',0)",
        [(f"M{i}", i, f"Modalidade {i}", "Objeto",
          float(100 - i), float(80 - i)) for i in range(1, 10)])
    db.commit()
    r = relatorios.gerar(db, "economia", {"ano": 2026}, "T", "SP", tmp_path)
    html = Path(r["html"]).read_text(encoding="utf-8")
    grafico = html.split("Por modalidade</h3>")[1].split("Por família de item")[0]
    for i in range(1, 10):
        assert f"Modalidade {i}" in grafico, f"Modalidade {i} sumiu do gráfico"


def test_economia_lista_fornecedores_com_documento_mascarado(db, tmp_path):
    """CNPJ e CPF convivem no `niFornecedor` do PNCP — a máscara é escolhida
    pelo número de dígitos (`documento()`), nunca aplicada às cegas."""
    db.executemany(
        "INSERT INTO itens (id, contratacao_controle, ano, descricao,"
        " fornecedor_ni, fornecedor_nome, valor_total_estimado,"
        " valor_total_homologado, referencia, raw)"
        " VALUES (?,?,2026,?,?,?,?,?,0,'{}')",
        [("A#1", "A", "MERENDA", "11222333000144", "ALIMENTOS SA",
          100.0, 80.0),
         ("A#2", "A", "CONSULTORIA", "12345678901", "JOSE DA SILVA",
          50.0, 45.0)])
    db.commit()
    r = relatorios.gerar(db, "economia", {"ano": 2026}, "T", "SP", tmp_path)
    html = Path(r["html"]).read_text(encoding="utf-8")
    assert "Economia por fornecedor" in html
    assert "11.222.333/0001-44" in html          # CNPJ, 14 dígitos
    assert "123.456.789-01" in html              # CPF, 11 dígitos
    # o de maior economia vem primeiro na tabela
    assert html.index("ALIMENTOS SA") < html.index("JOSE DA SILVA")


def test_sem_brasao_configurado_mantem_o_estandarte(db, tmp_path):
    r = relatorios.gerar(db, "contratacoes", {"ano": 2026}, "T", "SP", tmp_path)
    html = Path(r["html"]).read_text(encoding="utf-8")
    assert 'viewBox="0 0 64 64"' in html          # o estandarte, sem trocar
    assert "<img" not in html


def test_com_brasao_configurado_ele_substitui_o_estandarte(db, tmp_path):
    db.execute("INSERT INTO config (chave, valor) VALUES"
               " ('brasao', 'data:image/png;base64,QQ==')")
    db.commit()
    r = relatorios.gerar(db, "contratacoes", {"ano": 2026}, "T", "SP", tmp_path)
    html = Path(r["html"]).read_text(encoding="utf-8")
    assert ('<img src="data:image/png;base64,QQ==" alt="Brasão do '
            'município"') in html
    assert 'viewBox="0 0 64 64"' not in html      # estandarte saiu do cabeçalho
    assert "LICITARIVM" in html                   # mas segue no rodapé


def test_filtro_orgao_nos_relatorios(db, tmp_path):
    db.execute("UPDATE contratacoes SET orgao_cnpj='111'")
    db.execute("UPDATE contratacoes SET orgao_cnpj='222'"
               " WHERE numero_controle='A'")
    db.commit()
    assert relatorios.dados_contratacoes(db, ano=2026,
                                         orgao="222")["totais"]["n"] == 1
    assert relatorios.dados_executivo(db, 2026, orgao="222")["cards"]["n"] == 1
    r = relatorios.gerar(db, "contratacoes",
                         {"ano": 2026, "orgao": "222",
                          "orgao_nome": "Câmara de Testópolis"},
                         "Testópolis", "SP", tmp_path)
    assert "orgao_222" in r["html"]
    assert "Câmara de Testópolis" in Path(r["html"]).read_text(encoding="utf-8")








def test_num_contrato_normaliza():
    assert relatorios.num_contrato("0033/26", 2026) == "33/2026"
    assert relatorios.num_contrato("35", 2026) == "35/2026"
    assert relatorios.num_contrato("7/2026", 2026) == "7/2026"
    assert relatorios.num_contrato(None, 2026) is None
    assert relatorios.num_contrato("0042/25", None) == "42"


def test_fracionamento(db, tmp_path):
    db.execute("UPDATE contratacoes SET modalidade_id=8, unidade='Sec. Adm'"
               " WHERE ano=2026")
    db.commit()
    d = relatorios.dados_fracionamento(db, 2026, limites={"compras": 100})
    # A ("Merenda", homologado 80) e B ("Obras", sem homologado, cai no
    # estimado 200) são objetos sem relação — mesmo na mesma unidade, o
    # agrupamento por similaridade NÃO soma os dois (achado 2026-08-25:
    # antes, agrupar por `unidade` somava "Merenda" com "Obras" só por
    # serem da mesma secretaria). Total geral continua 280 (soma de TODAS
    # as dispensas, grupo à parte); o grupo maior sozinho é B, 200.
    assert d["n"] == 2 and d["total"] == 280.0
    assert d["unidades"][0]["pct"] == 200.0
    assert len(d["unidades"]) == 2
    r = relatorios.gerar(db, "fracionamento", {"ano": 2026},
                         "Testópolis", "SP", tmp_path)
    html = Path(r["html"]).read_text(encoding="utf-8")
    assert "Alerta de Fracionamento" in html and "autocontrole" in html
    assert r["xlsx"] and Path(r["xlsx"]).exists()


def test_fracionamento_tem_o_medidor_de_limite(db, tmp_path):
    """Pedido do usuário (2026-08-08): a tabela já tinha o farol em texto
    ("ACIMA DO LIMITE"/"Atenção") — o gráfico (porta de
    ui/painel.js:grafLimites) mostra a distância até lá num olhar só, com
    "×o limite" acima de 100% em vez de uma barra do tamanho da de 100%
    escondendo a gravidade."""
    db.execute("UPDATE contratacoes SET modalidade_id=8, unidade='Sec. Adm'"
               " WHERE ano=2026")
    db.commit()
    r = relatorios.gerar(db, "fracionamento", {"ano": 2026, "limites":
                         {"compras": 100}}, "T", "SP", tmp_path)
    html = Path(r["html"]).read_text(encoding="utf-8")
    assert html.count("<svg") >= 1
    assert "2,0× o limite" in html      # B ("Obras"): 200/100, maior grupo
    assert "var(--erro)" in html        # cor de estouro


def test_teto_da_dispensa_classifica_pelo_amparo_nao_pela_modalidade():
    """Achado 2026-08-12, portado do licitarium-relatorios: `modalidade_id=8`
    (Dispensa) não é sinônimo de "sujeita ao limite do art. 75, II" — o
    amparo (lido do `raw`) é quem decide, e três situações precisam de
    tratamento diferente."""
    # art. 75, II — compras/serviços comuns: responde ao teto de compras
    assert relatorios.teto_da_dispensa(
        "Lei 14.133/2021, Art. 75, II", 100.0, 200.0) == 100.0
    # formato curto (sem "Lei 14.133/2021,") também casa
    assert relatorios.teto_da_dispensa("Art. 75, II", 100.0, 200.0) == 100.0
    # art. 75, I — obras/serviços de engenharia: teto PRÓPRIO
    assert relatorios.teto_da_dispensa(
        "Lei 14.133/2021, Art. 75, I", 100.0, 200.0) == 200.0
    # inciso III (e demais) não tem teto por valor — "I" não pode casar
    # como prefixo de "III"
    assert relatorios.teto_da_dispensa(
        "Lei 14.133/2021, Art. 75, III", 100.0, 200.0) is None
    # outra lei (agricultura familiar etc.): dispensa própria, sem teto
    assert relatorios.teto_da_dispensa("Lei 11.947/2009", 100.0, 200.0) \
        is None
    # amparo ausente: conservador — não afirma limite que não se sabe
    assert relatorios.teto_da_dispensa(None, 100.0, 200.0) is None


def test_dispensa_sem_teto_por_valor_nao_soma_e_fica_declarada(db):
    """O falso positivo real que motivou a correção: R$ 1.057.448,50 de
    agricultura familiar (dispensa própria, Lei 11.947/2009, sem teto por
    valor) aparecia como "1688% do limite" — o maior valor da lista, sem
    irregularidade nenhuma."""
    raw_af = json.dumps({"amparoLegal": {"nome": "Lei 11.947/2009"}})
    db.execute(
        "INSERT INTO contratacoes (numero_controle, ano, sequencial,"
        " orgao_cnpj, unidade, modalidade_id, modalidade_nome, objeto,"
        " valor_homologado, data_publicacao, referencia, raw)"
        " VALUES ('AF1',2026,50,'111','Sec. Adm',8,'Dispensa',"
        " 'Gêneros da agricultura familiar',1057448.50,'2026-06-01',0,?)",
        (raw_af,))
    db.execute("UPDATE contratacoes SET modalidade_id=8, unidade='Sec. Adm'"
               " WHERE ano=2026 AND numero_controle IN ('A','B')")
    db.commit()

    d = relatorios.dados_fracionamento(db, 2026, limites={"compras": 100})
    # a dispensa sem teto não conta no total nem nas unidades...
    assert d["total"] == 280.0
    assert all(u["total"] != pytest.approx(1057448.50) for u in d["unidades"])
    # ...mas também não desaparece: sai declarada à parte
    assert len(d["fora_do_limite_legal"]) == 1
    assert d["fora_do_limite_legal"][0]["objeto"] == \
        "Gêneros da agricultura familiar"


def test_obra_e_medida_contra_o_limite_de_obras_nao_de_compras(db):
    """Art. 75, I (obras/serviços de engenharia) tem teto PRÓPRIO — antes,
    tudo caía no limite de compras e uma obra de R$ 150 mil marcava 150%
    quando o certo era medir contra o limite dobrado."""
    raw_obra = json.dumps({"amparoLegal": {"nome": "Lei 14.133/2021, Art. 75, I"}})
    db.execute(
        "INSERT INTO contratacoes (numero_controle, ano, sequencial,"
        " orgao_cnpj, unidade, modalidade_id, modalidade_nome, objeto,"
        " valor_homologado, data_publicacao, referencia, raw)"
        " VALUES ('OB1',2026,60,'111','Obras',8,'Dispensa',"
        " 'Reforma de telhado',150000.0,'2026-06-01',0,?)", (raw_obra,))
    db.commit()

    d = relatorios.dados_fracionamento(
        db, 2026, limites={"compras": 100000, "obras": 200000})
    obra = next(u for u in d["unidades"] if u["objeto"] == "Reforma de telhado")
    assert obra["tipo"] == "obras"
    assert obra["pct"] == pytest.approx(75.0)   # 150.000 / 200.000, não / 100.000


def test_similaridade_agrupa_objeto_parecido_mesmo_com_grafia_diferente(db):
    """Motor portado do SGCD (2026-08-25): Jaccard de tokens, não um radical
    fixo de N palavras — "AQUISIÇÃO DE PNEUS PARA VEÍCULOS" e "COMPRA DE
    PNEUS E CÂMARAS" têm overlap parcial ("pneus") mesmo sem bater as duas
    primeiras palavras."""
    raw = json.dumps({"amparoLegal": {"nome": "Art. 75, II"}})
    db.executemany(
        "INSERT INTO contratacoes (numero_controle, ano, sequencial,"
        " orgao_cnpj, orgao_nome, unidade, modalidade_id, modalidade_nome,"
        " objeto, valor_homologado, data_publicacao, referencia, raw)"
        " VALUES (?,2026,?,'111','PREF','Sec.',8,'Dispensa',?,"
        " 40000.0,'2026-06-01',0,?)",
        [("P1", 80, "PNEUS E CÂMARAS PARA VEÍCULOS", raw),
         ("P2", 81, "AQUISIÇÃO DE PNEUS E CÂMARAS DE VEÍCULOS", raw),
         # sem relação nenhuma com os pneus — não pode entrar no grupo só
         # por ser a mesma unidade/órgão
         ("P3", 82, "SERVIÇO DE JARDINAGEM", raw)])
    db.commit()

    d = relatorios.dados_fracionamento(db, 2026, limites={"compras": 100000})
    pneus = next(g for g in d["unidades"] if g["n"] == 2)
    assert pneus["total"] == pytest.approx(80000.0)
    assert {"P1", "P2"} == set(pneus["numeros_controle"])
    jardinagem = next(g for g in d["unidades"] if g["n"] == 1)
    assert jardinagem["numeros_controle"] == ["P3"]


def test_rotulo_do_grupo_corta_descricao_muito_longa(db):
    """Achado do usuário em PDF real (2026-08-28): o objeto do PNCP é a
    descrição inteira do edital, não um radical curto — sem corte, o
    gráfico (SVG do papel e ECharts da tela, os dois desenhados pra rótulo
    curto) empurrava a barra pra fora do cartão ou escondia o rótulo por
    completo. O rótulo agregado do grupo corta em 90 caracteres; a listagem
    "Dispensas do período" (por dispensa) continua com o texto inteiro."""
    longa = ("CONTRATAÇÃO DE EMPRESA PARA PRESTAÇÃO DE SERVIÇOS COMUNS DE "
             "APOIO TÉCNICO-CONTÁBIL, ASSISTÊNCIA E ORIENTAÇÃO OPERACIONAL "
             "AO DEPARTAMENTO DE CONTABILIDADE")
    assert len(longa) > 90   # a própria descrição de teste já estoura
    raw = json.dumps({"amparoLegal": {"nome": "Art. 75, II"}})
    db.execute(
        "INSERT INTO contratacoes (numero_controle, ano, sequencial,"
        " orgao_cnpj, orgao_nome, unidade, modalidade_id, modalidade_nome,"
        " objeto, valor_homologado, data_publicacao, referencia, raw)"
        " VALUES ('L1',2026,90,'111','PREF','Sec.',8,'Dispensa',?,"
        " 40000.0,'2026-06-01',0,?)", (longa, raw))
    db.commit()

    d = relatorios.dados_fracionamento(db, 2026, limites={"compras": 100000})
    grupo = d["unidades"][0]
    assert len(grupo["objeto"]) <= 90
    assert grupo["objeto"].endswith("…")
    # a dispensa individual (não o rótulo agregado) preserva o texto inteiro
    assert d["dispensas"][0]["objeto"] == longa


def test_janela_movel_ve_o_que_o_exercicio_esconde(db):
    """Achado do usuário (2026-08-25): fracionamento dividido na virada
    dez/jan não aparece no corte por exercício civil — cada dispensa cai
    num relatório de ano diferente. Período móvel (config `frac_janela`)
    resolve isso. Datas em dias corridos a partir de hoje (não meses de
    calendário) pra o teste não depender de em que mês do ano ele roda:
    400 dias é sempre > 12 meses e < 24 meses."""
    hoje = date.today()
    recente = (hoje - timedelta(days=20)).isoformat()
    antiga = (hoje - timedelta(days=400)).isoformat()
    raw = json.dumps({"amparoLegal": {"nome": "Art. 75, II"}})
    db.executemany(
        "INSERT INTO contratacoes (numero_controle, ano, sequencial,"
        " orgao_cnpj, orgao_nome, unidade, modalidade_id, modalidade_nome,"
        " objeto, valor_homologado, data_publicacao, referencia, raw)"
        " VALUES (?,?,?,'111','PREF','Sec.',8,'Dispensa','Material de"
        " limpeza',40000.0,?,0,?)",
        [("M1", int(recente[:4]), 90, recente, raw),
         ("M2", int(antiga[:4]), 91, antiga, raw)])
    db.commit()

    movel_12 = relatorios.dados_fracionamento(
        db, hoje.year, limites={"compras": 100000}, janela="12")
    assert movel_12["n"] == 1 and movel_12["total"] == pytest.approx(40000.0)

    movel_24 = relatorios.dados_fracionamento(
        db, hoje.year, limites={"compras": 100000}, janela="24")
    assert movel_24["n"] == 2 and movel_24["total"] == pytest.approx(80000.0)
    assert movel_24["janela"] == "24"

    exercicio = relatorios.dados_fracionamento(
        db, hoje.year, limites={"compras": 100000})
    assert exercicio["janela"] == "exercicio"


def test_dois_orgaos_sob_o_teto_nao_somam_pra_estourar(db):
    """O teto do art. 75 é "por órgão ou entidade" (§1º) — Prefeitura e
    Câmara que dispensam a mesma coisa têm cada uma o seu limite. Somar as
    duas contra um teto só acusaria fracionamento onde há duas compras
    legais."""
    raw_ii = json.dumps({"amparoLegal": {"nome": "Art. 75, II"}})
    db.executemany(
        "INSERT INTO contratacoes (numero_controle, ano, sequencial,"
        " orgao_cnpj, orgao_nome, unidade, modalidade_id, modalidade_nome,"
        " objeto, valor_homologado, data_publicacao, referencia, raw)"
        " VALUES (?,2026,?,?,?,'Sec. Adm',8,'Dispensa','Material',"
        " 60000.0,'2026-06-01',0,?)",
        [("PREF1", 70, "111", "PREFEITURA", raw_ii),
         ("CAM1", 71, "222", "CÂMARA", raw_ii)])
    db.commit()

    d = relatorios.dados_fracionamento(db, 2026, limites={"compras": 100000})
    pcts = {u["orgao_nome"]: u["pct"] for u in d["unidades"]
           if u["total"] == pytest.approx(60000.0)}
    assert pcts == {"PREFEITURA": pytest.approx(60.0),
                    "CÂMARA": pytest.approx(60.0)}
    # nenhuma das duas passa de 100% sozinha — juntas passariam (120%)
    assert all(pct < 100 for pct in pcts.values())


def test_render_fracionamento_mostra_tipo_e_orgao_quando_ha_mais_de_um(
        db, tmp_path):
    """Coluna Tipo (Obras/Compras) sempre aparece; coluna Órgão só quando
    há mais de um órgão com dispensa no exercício — mesmo padrão da
    coluna Unidade em Contratações."""
    raw_ii = json.dumps({"amparoLegal": {"nome": "Art. 75, II"}})
    db.executemany(
        "INSERT INTO contratacoes (numero_controle, ano, sequencial,"
        " orgao_cnpj, orgao_nome, unidade, modalidade_id, modalidade_nome,"
        " objeto, valor_homologado, data_publicacao, referencia, raw)"
        " VALUES (?,2026,?,?,?,'Sec. Adm',8,'Dispensa','Material',"
        " 1000.0,'2026-06-01',0,?)",
        [("PREF2", 80, "111", "PREFEITURA", raw_ii),
         ("CAM2", 81, "222", "CÂMARA", raw_ii)])
    db.commit()

    r = relatorios.gerar(db, "fracionamento", {"ano": 2026}, "T", "SP", tmp_path)
    html = Path(r["html"]).read_text(encoding="utf-8")
    assert "<th>Órgão</th>" in html
    assert "PREFEITURA" in html and "CÂMARA" in html
    assert "Compras" in html




def test_pagina_sem_categoria_fica_como_antes(db, tmp_path):
    """Painel e ficha de detalhe já têm identidade visual própria e chamam
    `_pagina` sem `categoria` — nenhuma etiqueta, régua ou faixa de acervo
    deve aparecer nesse caso, e o rodapé mantém o texto de sempre."""
    html = relatorios._pagina("Título", "<p>corpo</p>", "T", "SP", "período",
                              paisagem=False)
    assert 'class="etiqueta"' not in html
    assert 'class="regua"' not in html
    assert 'class="faixa-acervo"' not in html
    assert "Documento gerado automaticamente a partir de dados públicos" \
        in html


def test_gerar_traz_etiqueta_da_categoria_e_faixa_de_acervo(db, tmp_path):
    """`gerar()` computa o acervo uma vez (MAX(sync_em) entre as três
    tabelas) e passa a mesma fotografia — mesmo hash — pra qualquer tipo
    de relatório gerado a partir do mesmo estado do banco."""
    db.execute("UPDATE contratacoes SET sync_em='2026-08-13T10:22:00'")
    db.execute("UPDATE contratos SET sync_em='2026-08-13T10:22:00'")
    db.commit()

    r1 = relatorios.gerar(db, "contratacoes", {"ano": 2026}, "T", "SP", tmp_path)
    html1 = Path(r1["html"]).read_text(encoding="utf-8")
    assert '<span class="etiqueta">Cadastral</span>' in html1
    assert "Acervo sincronizado em 13/08/2026 ·" in html1
    assert "Apurado a partir do PNCP · acervo sincronizado em 13/08/2026" \
        in html1

    r2 = relatorios.gerar(db, "contratos", {"ano": 2026}, "T", "SP", tmp_path)
    html2 = Path(r2["html"]).read_text(encoding="utf-8")
    hash1 = html1.split("Acervo sincronizado em 13/08/2026 · ")[1][:6]
    hash2 = html2.split("Acervo sincronizado em 13/08/2026 · ")[1][:6]
    assert hash1 == hash2   # mesmo estado do banco, mesmo hash


def test_metodo_e_categoria_nos_cinco_relatorios_que_nao_tinham(db):
    """Achado 2026-08-13 (portado do diagnóstico de identidade): contratações,
    contratos, atas, executivo e economia não tinham nenhuma nota de método
    — só fracionamento, minuta e preços tinham. Os cinco ganham uma."""
    d_rel = relatorios.dados_contratacoes(db, ano=2026)
    assert '<div class="caixa-aviso">' in relatorios.render_contratacoes(
        d_rel, "T", "SP", "período", categoria=("Cadastral", "acento"))

    d_ct = relatorios.dados_contratos(db, ano=2026)
    assert '<div class="caixa-aviso">' in relatorios.render_contratos(
        d_ct, "T", "SP", "período")

    d_at = relatorios.dados_atas(db, ano=2026)
    assert '<div class="caixa-aviso">' in relatorios.render_atas(
        d_at, "T", "SP", "período")

    d_pn = relatorios.dados_painel(db, 2026)
    assert '<div class="caixa-aviso">' in relatorios.render_executivo(
        d_pn, "T", "SP")
    assert '<div class="caixa-aviso">' in relatorios.render_economia(
        d_pn, "T", "SP")


def test_documento_sai_com_a_paleta_institucional(db, tmp_path):
    """Documento oficial não tem tema (reversão consciente da v1.14.4).

    Lá o relatório passou a seguir o tema da tela porque forçava pergaminho
    e ignorava a escolha do usuário. Aqui a regra é mais forte: o papel que
    vai ao Tribunal de Contas é peça do município — branco, grafite, sem
    ouro nem vinho.

    Que ele não siga a tela virou garantia **estrutural** na v1.20.1: não
    existe mais parâmetro de tema em `gerar`/`render_*`/`_pagina` para
    seguir. O que sobra a testar é a paleta em si.
    """
    import inspect
    assert "tema" not in inspect.signature(relatorios.gerar).parameters

    r = relatorios.gerar(db, "contratacoes", {"ano": 2026}, "T", "SP", tmp_path)
    html = Path(r["html"]).read_text(encoding="utf-8")
    # a paleta é o que importa; o estandarte tem cores próprias cravadas no
    # SVG (design/IDENTIDADE.md: marca não troca de cor com a pele), então a
    # varredura é só no bloco de variáveis
    paleta = re.search(r":root \{(.*?)\}", html, re.S).group(1)
    assert "--bg:#ffffff" in paleta        # papel branco
    assert "#10151c" not in paleta         # nada do Observatório
    assert "#f5efe2" not in paleta         # nem do pergaminho
    assert "#8b2e2e" not in paleta         # nem o vinho do acento antigo
    assert "#b08d3e" not in paleta         # nem o dourado das réguas
    assert "double" not in html            # régua dupla de diploma saiu


def test_todos_os_relatorios_saem_em_paisagem(db, tmp_path):
    """Pedido do usuário (2026-08-08): melhor uso da largura da página —
    executivo e fracionamento eram os dois únicos ainda em retrato."""
    db.execute("UPDATE contratacoes SET modalidade_id=8 WHERE ano=2026")
    db.commit()
    for tipo, params in (("executivo", {"ano": 2026}),
                         ("fracionamento", {"ano": 2026}),
                         ("contratacoes", {"ano": 2026}),
                         ("contratos", {}), ("atas", {})):
        r = relatorios.gerar(db, tipo, params, "T", "SP", tmp_path)
        html = Path(r["html"]).read_text(encoding="utf-8")
        assert "landscape" in html, f"{tipo} não saiu em paisagem"
        assert "portrait" not in html, f"{tipo} ainda tem @page portrait"


def test_executivo_usa_os_graficos_do_painel(db, tmp_path):
    """Pedido do usuário (2026-08-08): o resumo executivo reaproveita os
    gráficos do Painel (mesma consulta, dados_painel) em vez de só tabelas
    com uma barra de largura fixa via CSS."""
    r = relatorios.gerar(db, "executivo", {"ano": 2026}, "T", "SP", tmp_path)
    html = Path(r["html"]).read_text(encoding="utf-8")
    assert html.count("<svg") >= 2       # sparkline do hero + colunas do mês
    assert 'class="card hero"' in html
    assert 'class="card kpiv"' in html
    assert "Por modalidade — valor homologado" in html


def test_minuta_pca_mostra_a_curva_abc():
    """Pedido do usuário (2026-08-08): pca_builder.classificar_abc já rodava
    dentro de listar_minuta e alimentava a tela de Montar PCA, mas a classe
    nunca aparecia no documento impresso — só a lista crua de itens."""
    d = {"ano": 2027, "parametros": {"margem": 10},
         "totais": {"grupos": 3, "valor": 1000.0},
         "itens": [
             {"descricao": "Item A", "categoria": "Material", "unidade": "UN",
              "quantidade": 1, "valor_unitario": 800.0, "valor_total": 800.0,
              "abc": "A"},
             {"descricao": "Item B", "categoria": "Material", "unidade": "UN",
              "quantidade": 1, "valor_unitario": 150.0, "valor_total": 150.0,
              "abc": "B"},
             {"descricao": "Item C", "categoria": "Material", "unidade": "UN",
              "quantidade": 1, "valor_unitario": 50.0, "valor_total": 50.0,
              "abc": "C"},
         ]}
    html = relatorios.render_minuta_pca(d, "T", "SP")
    assert "Curva ABC" in html
    assert "1 item classe A = 80% do valor" in html
    assert "1 item classe B = 15% do valor" in html
    assert "1 item classe C = 5% do valor" in html
    assert '<th class="ctr" title="Curva ABC' in html


def test_tipo_desconhecido(db, tmp_path):
    with pytest.raises(ValueError):
        relatorios.gerar(db, "xxx", {}, "T", "SP", tmp_path)


def test_documento_distingue_cnpj_de_cpf():
    """O `niFornecedor` do PNCP guarda os dois — no acervo real há 34 CPFs.

    Máscara de CNPJ aplicada às cegas transformaria 01472188616 em
    "01.472.188/616-" e o relatório sairia com o documento adulterado.
    """
    assert relatorios.documento("13286494000164") == "13.286.494/0001-64"
    assert relatorios.documento("01472188616") == "014.721.886-16"
    # já formatado na origem continua correto (idempotente)
    assert relatorios.documento("13.286.494/0001-64") == "13.286.494/0001-64"
    # o que não é nenhum dos dois sai como veio, sem inventar pontuação
    assert relatorios.documento("A1B2") == "A1B2"
    assert relatorios.documento("123") == "123"
    assert relatorios.documento(None) == "–"
    assert relatorios.documento("") == "–"


def test_relatorios_imprimem_documento_com_mascara(db):
    db.execute(
        "INSERT INTO contratos (numero_controle, orgao_cnpj, fornecedor_ni,"
        " fornecedor_nome, objeto, valor_global, data_publicacao, raw)"
        " VALUES ('K-1','111','13286494000164','FORN LTDA','Objeto',10.0,"
        " '2026-02-02','{}')")
    db.execute(
        "INSERT INTO contratos (numero_controle, orgao_cnpj, fornecedor_ni,"
        " fornecedor_nome, objeto, valor_global, data_publicacao, raw)"
        " VALUES ('K-2','111','01472188616','JOSE DA SILVA','Objeto',10.0,"
        " '2026-02-03','{}')")
    db.commit()
    html = relatorios.render_contratos(
        relatorios.dados_contratos(db, ano=2026), "Orindiúva", "SP", "2026")
    assert "13.286.494/0001-64" in html
    assert "014.721.886-16" in html
    assert "13286494000164" not in html      # nada de número cru


def test_url_pncp_do_processo():
    assert relatorios.url_pncp("96291141000180", 2024, 4344) == \
        "https://pncp.gov.br/app/editais/96291141000180/2024/4344"
    # sem os três dados não há link: melhor nenhum do que quebrado
    for faltando in (("", 2024, 1), ("111", None, 1), ("111", 2024, None)):
        assert relatorios.url_pncp(*faltando) is None




# ── auditoria de segurança (2026-08-09): dois sinks de XSS armazenado ────────
# O relatório é aberto com `webbrowser.open` no navegador REAL do usuário
# (origem file://), não dentro do WebView — script que sai daqui executa fora
# da janela do programa. Vetor confirmado: `importar_acervo` troca o banco
# inteiro por um .zip de terceiro, validado só por `quick_check`, sem
# conferência de tipo de coluna.



def test_moeda_nao_numerica_nao_derruba_o_relatorio():
    """moeda()/moeda_fina() não tinham a mesma proteção de quantidade() —
    TEXT numa coluna REAL (auditoria 2026-08-09/11) fazia f"{v:,.2f}"
    levantar ValueError e derrubar o relatório inteiro, não só aquela
    célula."""
    assert relatorios.moeda("N/D") == "–"
    assert relatorios.moeda(None) == "–"
    assert relatorios.moeda(100.0) == "R$ 100,00"
    assert relatorios.moeda_fina("N/D") == "–"
    assert relatorios.moeda_fina(None) == "–"
    assert relatorios.moeda_fina(0.0466) == "R$ 0,0466"












def test_categoria_relatorio_cobre_os_sete_tipos_em_quatro_cores():
    """Selo de procedência: cada tipo de relatório tem categoria e cor
    (Cadastral/Analítico/Vigilância/Planejamento). A parte de preços saiu do
    Free (virou produto à parte)."""
    tipos = {"contratacoes", "contratos", "atas", "executivo", "economia",
             "fracionamento", "minuta_pca"}
    assert set(relatorios.CATEGORIA_RELATORIO) == tipos
    assert "precos" not in relatorios.CATEGORIA_RELATORIO
    assert "comparados" not in relatorios.CATEGORIA_RELATORIO
    cores = {cor for _, cor in relatorios.CATEGORIA_RELATORIO.values()}
    assert len(cores) == 4
    rotulos = {r for r, _ in relatorios.CATEGORIA_RELATORIO.values()}
    assert rotulos == {"Cadastral", "Analítico", "Vigilância", "Planejamento"}
