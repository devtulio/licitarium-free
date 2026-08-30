"""Exportação da lista em planilha (.xlsx) — substituiu o CSV cru
(pedido do usuário, 2026-08-29): colunas curadas (sem `raw`/`sync_em`),
cabeçalho traduzido.
"""
import json
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
import licitarium


class JanelaFalsa:
    def __init__(self, resposta=None):
        self.resposta = resposta

    def create_file_dialog(self, tipo, **kwargs):
        return self.resposta


@pytest.fixture
def api(tmp_path, monkeypatch):
    monkeypatch.setattr(licitarium, "DIR_DADOS", tmp_path)
    monkeypatch.setattr(licitarium, "ARQUIVO_DB", tmp_path / "t.db")
    db = licitarium.abrir_db()
    raw = json.dumps({"amparoLegal": {"nome": "Art. 75, II"}})
    db.execute(
        "INSERT INTO contratacoes (numero_controle, ano, sequencial,"
        " modalidade_nome, orgao_nome, unidade, objeto, situacao,"
        " valor_estimado, valor_homologado, data_publicacao, raw, sync_em)"
        " VALUES ('C1',2026,10,'Dispensa','Prefeitura','Sec. Adm',"
        " 'Aquisição de merenda escolar','Homologada',1000.0,900.0,"
        " '2026-03-01',?,'2026-03-02T10:00:00')", (raw,))
    db.commit()
    db.close()
    api = licitarium.Api()
    api._janela = JanelaFalsa()
    return api


def test_exportar_lista_cura_colunas_e_traduz_cabecalho(api, tmp_path):
    destino = tmp_path / "contratacoes.xlsx"
    api._janela.resposta = str(destino)

    r = api.exportar_planilha("contratacoes", {})
    assert r["ok"] and r["linhas"] == 1

    ws = openpyxl.load_workbook(destino).active
    cabecalho = [c.value for c in ws[1]]
    # colunas técnicas/internas não vazam pra planilha do usuário
    assert "raw" not in cabecalho and "sync_em" not in cabecalho
    assert "OBJETO" in cabecalho and "SITUAÇÃO" in cabecalho
    assert "VALOR HOMOLOGADO" in cabecalho
    # par estimado/homologado presente → coluna Deságio por fórmula
    assert "DESÁGIO" in cabecalho
    linha = dict(zip(cabecalho, [c.value for c in ws[2]]))
    assert linha["OBJETO"] == "AQUISIÇÃO DE MERENDA ESCOLAR"  # sempre caixa alta
    assert linha["VALOR HOMOLOGADO"] == 900.0


def test_exportar_contratos_numero_limpo_e_uma_linha_por_fornecedor_na_ata(
        api, tmp_path):
    """Achados do usuário em screenshot (2026-08-30): (1) número do
    contrato saía cru do PNCP ("0049/26") — vira sequencial/ano sem zero à
    esquerda ("49/2026"); (2) ata com mais de um fornecedor vencedor (RP de
    vários itens) vira uma linha por fornecedor, como Contratos já tem."""
    db = licitarium.abrir_db()
    db.execute(
        "INSERT INTO contratos (numero_controle, orgao_cnpj, numero_contrato,"
        " ano_contrato, sequencial_contrato, fornecedor_ni, fornecedor_nome,"
        " objeto, valor_global, vigencia_inicio, vigencia_fim, data_publicacao)"
        " VALUES ('CT1','111','0049/26',2026,49,'999','FORN X',"
        " 'objeto do contrato',5000.0,'2026-08-21','2026-10-21','2026-08-26')")
    db.execute(
        "INSERT INTO atas (numero_controle, contratacao_controle, orgao_cnpj,"
        " numero_ata, ano_ata, objeto, fornecedor_ni, fornecedor_nome,"
        " vigencia_inicio, vigencia_fim)"
        " VALUES ('AT1','C1','111','12',2026,'objeto da ata',"
        " '111\x1f222','FORN A\x1fFORN B','2026-01-01','2027-01-01')")
    db.commit()
    db.close()

    destino = tmp_path / "contratos.xlsx"
    api._janela.resposta = str(destino)
    r = api.exportar_planilha("contratos", {})
    assert r["ok"]
    ws = openpyxl.load_workbook(destino).active
    cabecalho = [c.value for c in ws[1]]
    linha = dict(zip(cabecalho, [c.value for c in ws[2]]))
    assert linha["NUMERO CONTRATO"] == "49/2026"
    assert "VENCIMENTO (DIAS)" in cabecalho  # coluna nova, por fórmula

    destino_atas = tmp_path / "atas.xlsx"
    api._janela.resposta = str(destino_atas)
    r = api.exportar_planilha("atas", {})
    assert r["ok"] and r["linhas"] == 2   # 1 ata × 2 fornecedores
    ws = openpyxl.load_workbook(destino_atas).active
    assert ws.max_row == 3                # cabeçalho + 2 linhas (1 por fornecedor)
    cabecalho = [c.value for c in ws[1]]
    linha1 = dict(zip(cabecalho, [c.value for c in ws[2]]))
    linha2 = dict(zip(cabecalho, [c.value for c in ws[3]]))
    # CNPJ/CPF vira número de verdade (leva máscara própria — não dá pra
    # mascarar número em texto), pedido do usuário (2026-08-30)
    assert {linha1["CNPJ/CPF DO FORNECEDOR"], linha2["CNPJ/CPF DO FORNECEDOR"]} \
        == {111, 222}
