"""Relatórios do Licitarium — relações oficiais (TCE) e resumo executivo.

Gera HTML standalone timbrado (imprimível pelo navegador, título vira nome do
PDF) e planilha .xlsx para as relações. Só stdlib, com uma exceção: a
exportação em planilha usa `openpyxl` (pura Python, sem dependência nativa),
importado só dentro de `escrever_planilha` — o resto do módulo continua sem
precisar dela.
"""
import hashlib
import html
import json
import math
import re
import unicodedata
from collections import Counter
from datetime import date, datetime

import pca_builder

# a arte vem de marca.py, gerado por design/gerar_marca.py a partir do
# design/estandarte-t3.svg — antes era uma cópia mantida à mão aqui
from marca import ESTANDARTE

TITULOS = {"contratacoes": "Relação de Contratações",
           "contratos": "Relação de Contratos",
           "atas": "Relação de Atas de Registro de Preços",
           "executivo": "Resumo Executivo de Contratações",
           "economia": "Economia e Comparativos por Modalidade e Categoria",
           "fracionamento": "Alerta de Fracionamento — Dispensas × Limites",
           "minuta_pca": "Minuta do Plano de Contratações Anual"}

# Valores do art. 75, I e II, da Lei 14.133/2021 conforme Decreto de
# atualização — parametrizáveis nas configurações (confira o decreto vigente).
# ESPELHO do Licitarium Pro (licitarium/servicos/consultas/painel.py): a
# portaria de atualização muda estes valores nos DOIS sistemas.
LIMITE_PADRAO_OBRAS = 125279.84
LIMITE_PADRAO_COMPRAS = 62639.92


def _inciso_dispensa(amparo):
    """"I", "II" ou `None` — qual inciso do art. 75 ampara a dispensa, lido
    do amparo legal bruto (não é campo da tabela). Casa em qualquer ponto da
    string (com ou sem o "Lei 14.133/2021," na frente — o PNCP varia a
    formatação); o `\\b` depois do algarismo garante que "III" (ou incisos
    maiores) nunca casa como "II"."""
    if not amparo:
        return None
    m = re.search(r"Art\.\s*75,\s*(I{1,2})\b", amparo)
    return m.group(1) if m else None


def teto_da_dispensa(amparo, limite_compras, limite_obras):
    """Qual teto de valor se aplica a uma dispensa — ou `None` quando não há.

    `modalidade_id=8` (Dispensa) NÃO é sinônimo de "sujeita ao limite do
    art. 75, II" — o amparo legal distingue três situações:

    - **Art. 75, II** — compras e serviços comuns: é o único que responde
      ao limite de compras.
    - **Art. 75, I** — obras e serviços de engenharia: teto PRÓPRIO
      (`LIMITE_PADRAO_OBRAS`, o dobro do de compras).
    - **Demais incisos** (emergência, licitação deserta, agricultura
      familiar/Lei 11.947 etc.) e **outras leis**: a dispensa vem da
      natureza do objeto, não do valor — não há teto contra o qual
      comparar.

    Achado 2026-08-12, portado da varredura do licitarium-relatorios: uma
    compra de R$ 1.057.448,50 de gêneros da agricultura familiar (Lei
    11.947/2009, dispensa PRÓPRIA, sem teto por valor) aparecia como
    "1688% do limite do art. 75, II" — falso positivo grave, acusação
    sobre o maior valor da lista onde não havia irregularidade nenhuma.

    Amparo ausente é tratado como "sem teto" — conservador: sem saber o
    amparo não se afirma que um limite se aplica.
    """
    inciso = _inciso_dispensa(amparo)
    if inciso is None:
        return None
    return limite_compras if inciso == "II" else limite_obras


# ── agrupamento por similaridade (motor portado do SGCD, 2026-08-25) ───────
# Antes, dispensas do mesmo órgão/teto entravam no mesmo grupo por baterem a
# mesma `unidade` do PNCP (que costuma trazer só o nome do órgão — no
# acervo do piloto, TODAS caem em "MUNICIPIO DE ORINDIUVA") ou o mesmo
# radical fixo de 2 palavras da descrição. As duas somavam coisas sem
# relação ("Merenda" e "Reforma de telhado" da mesma secretaria) e perdiam
# coisas relacionadas com grafia diferente ("PNEUS PARA VEÍCULOS" x "PNEUS E
# CÂMARAS"). O SGCD (sistema irmão, workflow de dispensa de um órgão só)
# já tinha resolvido isso com similaridade textual — portado aqui.
_FRAC_STOPWORDS = {"de", "da", "do", "das", "dos", "e", "em", "para", "com",
                   "a", "o", "as", "os", "no", "na", "nos", "nas", "pela",
                   "pelo", "por", "sem", "sob", "sobre", "um", "uma"}


def _frac_tokenize(texto):
    """Tokens de 3+ letras, sem acento/pontuação/palavra vazia."""
    limpo = unicodedata.normalize("NFD", (texto or "")) \
        .encode("ascii", "ignore").decode("ascii").lower()
    limpo = re.sub(r"[^a-z0-9\s]", " ", limpo)
    return {t for t in limpo.split() if len(t) > 2 and t not in _FRAC_STOPWORDS}


def _frac_similaridade(a, b):
    """Jaccard dos tokens de dois objetos: 0 (nada em comum) a 1 (mesmos
    tokens depois de normalizar)."""
    ta, tb = _frac_tokenize(a), _frac_tokenize(b)
    if not ta or not tb:
        return 0.0
    inter = len(ta & tb)
    return inter / (len(ta) + len(tb) - inter)


def _frac_mesma_unidade(a, b):
    ua, ub = (a or "").strip().lower(), (b or "").strip().lower()
    if not ua or not ub:
        return False
    return ua == ub or ua in ub or ub in ua


def _agrupar_por_similaridade(linhas):
    """Une dispensas do MESMO órgão e MESMO teto (`_teto` já anotado em cada
    linha) num grupo quando o objeto é parecido: Jaccard > 45% sozinho, ou
    > 20% quando também é a mesma unidade requisitante — o texto fraco
    ganha força quando o contexto organizacional bate junto.

    União por componente conexo (union-find): se A~B e B~C, os três entram
    no mesmo grupo mesmo que A e C sozinhos não passem no limiar — mesma
    regra do motor original (SGCD `analisarFracionamento`).
    """
    n = len(linhas)
    pai = list(range(n))

    def acha(i):
        while pai[i] != i:
            pai[i] = pai[pai[i]]
            i = pai[i]
        return i

    def une(i, j):
        ri, rj = acha(i), acha(j)
        if ri != rj:
            pai[ri] = rj

    for i in range(n):
        for j in range(i + 1, n):
            a, b = linhas[i], linhas[j]
            if a["orgao_cnpj"] != b["orgao_cnpj"] or a["_teto"] != b["_teto"]:
                continue
            sim = _frac_similaridade(a["objeto"], b["objeto"])
            mesma_unidade = _frac_mesma_unidade(a["unidade"], b["unidade"])
            if sim > 0.45 or (sim > 0.20 and mesma_unidade):
                une(i, j)

    grupos = {}
    for i, l in enumerate(linhas):
        grupos.setdefault(acha(i), []).append(l)
    return list(grupos.values())


def _linha_do_grupo(grupo):
    total = sum(l["valor"] or 0 for l in grupo)
    principal = max(grupo, key=lambda l: l["valor"] or 0)
    limite = principal["_teto"]
    outros = len({l["objeto"] for l in grupo if l["objeto"]}) - 1
    texto = (principal["objeto"] or "(sem descrição)").strip()
    # o objeto do PNCP costuma ser a descrição inteira do edital (não um
    # radical curto) — sem teto, o gauge (SVG do papel e ECharts da tela,
    # os dois desenhados pra rótulo curto) empurrava a barra pra fora do
    # cartão ou escondia o rótulo (achado do usuário, 2026-08-28, PDF real).
    # A listagem "Dispensas do período" (por dispensa, não agrupada) segue
    # com o objeto completo — só este rótulo agregado corta.
    if len(texto) > 90:
        texto = texto[:87].rstrip() + "…"
    rotulo = texto
    if outros > 0:
        rotulo += (f" (+{outros} objeto{'s' if outros > 1 else ''}"
                   f" semelhante{'s' if outros > 1 else ''})")
    return {"objeto": rotulo, "orgao_nome": principal["orgao_nome"],
            "n": len(grupo), "total": total, "limite": limite,
            "tipo": principal["_tipo"],
            "pct": total / limite * 100 if limite else 0,
            # o clique no alerta do Painel filtra por ESTES processos — não
            # dá mais pra recalcular o grupo em SQL (não é mais um radical
            # fixo de N palavras), então a lista de membros vai junto
            "numeros_controle": [l["numero_controle"] for l in grupo]}


# Janela de análise, configurável (Configurações → Limites de dispensa):
# "exercicio" (padrão — ano civil, é o que TCE/AGU usam, Lei 4.320/64 art.
# 34) ou N meses corridos até hoje (período móvel — mais rigoroso, pega
# dispensa dividida na virada dez/jan, que o corte por exercício não vê).
JANELAS_FRAC_MESES = {"12": 12, "18": 18, "24": 24}

MESES_NOME = ["jan", "fev", "mar", "abr", "mai", "jun",
              "jul", "ago", "set", "out", "nov", "dez"]

# O PNCP preenche `categoria` com "Não se aplica" na quase totalidade dos
# itens — a string é truthy, então um `COALESCE`/`or` simples nunca cai pro
# fallback e "Economia por categoria" sai com UMA barra sem informação
# nenhuma. `material_servico` tem conteúdo de verdade. Achado 2026-08-12,
# portado da varredura do licitarium-relatorios (Django).
CATEGORIA_VAZIA = {"não se aplica", "nao se aplica",
                   "não informado", "nao informado"}


def _categoria_util(categoria, material_servico):
    if (categoria or "").strip().lower() in CATEGORIA_VAZIA:
        categoria = None
    return categoria or material_servico or "(sem categoria)"


def _e(v):
    return html.escape(str(v)) if v is not None else "–"


def documento(v):
    """Formata o identificador do fornecedor conforme o que ele é.

    O campo do PNCP (`niFornecedor`) guarda CNPJ e também CPF — no acervo
    real são 14 registros de pessoa física em contratos e 20 em itens.
    Máscara de CNPJ aplicada às cegas estragaria justamente esses.
    """
    digitos = re.sub(r"\D", "", str(v or ""))
    if len(digitos) == 14:
        return (f"{digitos[:2]}.{digitos[2:5]}.{digitos[5:8]}"
                f"/{digitos[8:12]}-{digitos[12:]}")
    if len(digitos) == 11:
        return f"{digitos[:3]}.{digitos[3:6]}.{digitos[6:9]}-{digitos[9:]}"
    return str(v) if v else "–"    # identificador estrangeiro ou ausente


def moeda(v):
    """Mesma proteção de quantidade(): a coluna promete REAL, mas afinidade
    do SQLite não converte texto não-numérico — fica gravado como TEXT."""
    try:
        v = float(v)
    except (TypeError, ValueError):
        return "–"
    inteiro, decimal = f"{v:,.2f}".split(".")
    return "R$ " + inteiro.replace(",", ".") + "," + decimal


def moeda_fina(v):
    """Preço de unidade-base tem centavo de centavo: R$ 0,0466 por folha."""
    try:
        v = float(v)
    except (TypeError, ValueError):
        return "–"
    if v >= 1:
        return moeda(v)
    inteiro, decimal = f"{v:,.4f}".split(".")
    return "R$ " + inteiro.replace(",", ".") + "," + decimal


def quantidade(v):
    """Quantidade na coluna numérica: número formatado, ou travessão.

    A coluna `itens.quantidade_homologada` é REAL, mas afinidade do SQLite
    **não converte** texto não-numérico — ele fica gravado como TEXT e
    chegava cru ao HTML do relatório, que é aberto no navegador real do
    usuário (achado da auditoria de segurança, 2026-08-09). A coluna
    promete número; o que não for número não é exibido.
    """
    try:
        n = float(v)
    except (TypeError, ValueError):
        return "–"
    if n == int(n):
        return f"{int(n):,}".replace(",", ".")
    inteiro, decimal = f"{n:,.2f}".split(".")
    return inteiro.replace(",", ".") + "," + decimal


def compacto(v):
    """Número curto pra rótulo de gráfico — mesma régua de ui/painel.js."""
    if v is None:
        return "–"
    av = abs(v)
    if av >= 1e6:
        return f"R$ {v / 1e6:.1f}".replace(".", ",") + " mi"
    if av >= 1e3:
        return f"R$ {v / 1e3:.0f} mil"
    return moeda(v)


def _escala(maximo):
    """Eixo com números redondos (1/2/2,5/5×10^n) — nunca '1,7 mi' e '3,3 mi'."""
    if not maximo or maximo <= 0:
        return 1, 0.25
    p = 10 ** math.floor(math.log10(maximo / 3))
    for m in (1, 2, 2.5, 5, 10):
        passo = m * p
        if maximo / passo <= 4.2:
            return math.ceil(maximo / passo) * passo, passo
    return maximo, maximo / 4


def _svg(largura, altura, dentro):
    return (f'<svg viewBox="0 0 {largura} {altura}" width="100%"'
            f' height="{altura}" role="img"'
            f' preserveAspectRatio="xMidYMid meet">{dentro}</svg>')


def _grafico_meses(meses, cor, larg=900):
    """Colunas pareadas estimado (claro) × homologado (cheio) — porta de
    ui/painel.js:grafMeses, mesma leitura na tela e no papel."""
    if not any(m["valor"] or m["estimado"] for m in meses):
        return '<div class="vazio">Sem contratações no exercício.</div>'
    ultimo = 0
    for i, m in enumerate(meses):
        if m["valor"] or m["estimado"]:
            ultimo = i
    dados = meses[:max(ultimo + 1, date.today().month)]
    alto = base = 170
    topo, passo = _escala(max((max(m["valor"], m["estimado"]) for m in dados),
                              default=0))

    def y(v):
        return base - (v / topo) * (base - 30) if topo else base

    passo_x = (larg - 60) / len(dados)
    g = ""
    v = 0.0
    while v <= topo + 1e-6:
        g += (f'<line class="eixo" x1="48" y1="{y(v):.1f}" x2="{larg - 8}"'
              f' y2="{y(v):.1f}" opacity="{1 if not v else .55}"/>'
              f'<text class="rot" x="44" y="{y(v) + 4:.1f}" text-anchor="end">'
              f'{"0" if not v else compacto(v).replace("R$ ", "")}</text>')
        v += passo
    for i, m in enumerate(dados):
        x = 56 + i * passo_x
        w = min(34, passo_x / 2.6)
        he = max(2, base - y(m["estimado"]))
        hh = max(2, base - y(m["valor"]))
        g += (f'<rect x="{x:.1f}" y="{y(m["estimado"]):.1f}" width="{w:.1f}"'
              f' height="{he:.1f}" rx="4" fill="{cor}" opacity=".32"/>'
              f'<rect x="{x + w + 2:.1f}" y="{y(m["valor"]):.1f}" width="{w:.1f}"'
              f' height="{hh:.1f}" rx="4" fill="{cor}"/>'
              f'<text class="rot" x="{x + w + 1:.1f}" y="{base + 16}"'
              f' text-anchor="middle">{MESES_NOME[m["mes"] - 1]}</text>')
    legenda = (f'<div class="leg"><span><i style="background:{cor};'
              f'opacity:.32"></i>Estimado</span>'
              f'<span><i style="background:{cor}"></i>Homologado</span></div>')
    return _svg(larg, alto + 26, g) + legenda


def _grafico_barras(itens, valor, rotulo, cor, larg=900, sub=None):
    """Barras horizontais, uma série, rótulo direto — porta de
    ui/painel.js:grafBarras."""
    if not itens:
        return '<div class="vazio">Sem dados no exercício.</div>'
    maximo = max(valor(it) for it in itens) or 1
    linha = 40
    g = ""
    for i, it in enumerate(itens):
        y = i * linha + 18
        w = max(3, (valor(it) / maximo) * (larg - 110))
        extra = f" · {_e(sub(it))}" if sub else ""
        g += (f'<text class="rot" x="0" y="{y - 6}">{_e(rotulo(it))}{extra}</text>'
              f'<rect x="0" y="{y}" width="{w:.1f}" height="17" rx="4"'
              f' fill="{cor}"/>'
              f'<text class="val" x="{w + 8:.1f}" y="{y + 14}">'
              f'{compacto(valor(it))}</text>')
    return _svg(larg, len(itens) * linha + 6, g)


def _grafico_limites(unidades, larg=760):
    """Medidor por unidade × limite do art. 75 — porta de
    ui/painel.js:grafLimites. A barra cheia diz "chegou ao limite"; passar
    dele é gravidade diferente, e "874%" numa barra do tamanho da de 100%
    esconderia isso — por isso o texto vira "×o limite" acima de 100%."""
    if not unidades:
        return '<div class="vazio">Nenhuma dispensa registrada no período.</div>'
    bloco = 66
    g = ""
    for i, u in enumerate(unidades):
        y = i * bloco + 16
        cheio = larg - 60
        pct = u.get("pct") or 0
        w = min(1, pct / 100) * cheio
        estourou = pct > 100
        cor = ("var(--erro)" if pct >= 90 else
               "var(--warn)" if pct >= 75 else "var(--s3)")
        texto_pct = (f"{pct / 100:.1f}".replace(".", ",") + "× o limite"
                     if estourou else f"{pct:.0f}% do limite")
        g += (f'<text class="rot" x="0" y="{y - 4}">{_e(u["objeto"])} · '
              f'{u["n"]} {"dispensa" if u["n"] == 1 else "dispensas"}</text>'
              f'<rect x="0" y="{y}" width="{cheio:.1f}" height="14" rx="4"'
              f' fill="var(--surface2)"/>'
              f'<rect x="0" y="{y}" width="{max(3, w):.1f}" height="14" rx="4"'
              f' fill="{cor}"/>')
        if estourou:
            g += (f'<path d="M{cheio - 1:.1f},{y - 3} l10,10 l-10,10 z"'
                  f' fill="var(--erro)"/>')
        g += (f'<text class="val" x="0" y="{y + 30}">{moeda(u["total"])} · '
              f'<tspan fill="{cor}" font-weight="600">{texto_pct}</tspan>'
              f'</text>')
    return _svg(larg, len(unidades) * bloco + 6, g)




def url_pncp(cnpj, ano, sequencial):
    """Página do processo no portal — a mesma que o programa abre na tela.

    No relatório serve à transparência: quem recebe o documento confere cada
    preço na fonte oficial, em vez de confiar na nossa tabela.
    """
    if not (cnpj and ano and sequencial):
        return None
    return f"https://pncp.gov.br/app/editais/{cnpj}/{ano}/{sequencial}"


def num_contrato(numero, ano):
    """PNCP grava '0033/26'; padrão de exibição é numero/ano: 33/2026."""
    if not numero:
        return None
    m = re.match(r"0*(\d+)", str(numero))
    n = m.group(1) if m else str(numero)
    return f"{n}/{ano}" if ano else str(n)


def data_br(s):
    if not s:
        return "–"
    p = str(s)[:10].split("-")
    return f"{p[2]}/{p[1]}/{p[0]}" if len(p) == 3 else str(s)


# ── exportação em planilha ──────────────────────────────────────────────────
# Rótulo amigável pro cabeçalho — a mesma chave técnica (coluna do SQL/campo
# dos dicts de dados_*) aparece em vários relatórios e no export da lista, daí
# um dicionário só. O que não está aqui vira Title Case da própria chave
# (`_rotulo_campo`), sem quebrar — só menos bonito.
ROTULOS_EXPORT = {
    "numero_controle": "Número de controle", "numero": "Número",
    "sequencial": "Sequencial", "ano": "Ano",
    "ano_contrato": "Ano do contrato", "ano_ata": "Ano da ata",
    "modalidade_nome": "Modalidade", "amparo": "Amparo legal",
    "objeto": "Objeto", "unidade": "Unidade",
    "orgao_cnpj": "CNPJ do órgão", "orgao_nome": "Órgão",
    "fornecedor_nome": "Fornecedor", "fornecedor_ni": "CNPJ/CPF do fornecedor",
    "valor_estimado": "Valor estimado", "valor_homologado": "Valor homologado",
    "valor_global": "Valor", "valor_total": "Valor total",
    "valor_unitario": "Valor unitário", "quantidade": "Quantidade",
    "margem": "Margem", "categoria": "Categoria", "descricao": "Descrição",
    "grupo": "Grupo", "situacao": "Situação",
    "contratacao_controle": "Contratação de origem", "numero_item": "Item",
    "data_publicacao": "Publicação",
    "data_encerramento_proposta": "Encerramento da proposta",
    "vigencia_inicio": "Vigência inicial", "vigencia_fim": "Vigência final",
}


def _rotulo_campo(chave):
    return ROTULOS_EXPORT.get(chave) or " ".join(
        p.capitalize() for p in chave.split("_"))


def escrever_planilha(caminho, linhas):
    """Grava `linhas` (lista de dicts, mesma chave em todas) num .xlsx
    legível: cabeçalho traduzido e destacado, congelado no topo, largura de
    coluna pelo conteúdo, número com separador de milhar. Substitui o CSV
    cru — pedido do usuário (2026-08-29), abre "bonitinho" no Excel/
    LibreOffice sem precisar reformatar nada.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    chaves = list(linhas[0].keys())
    ws.append([_rotulo_campo(k) for k in chaves])
    for cel in ws[1]:
        cel.font = Font(bold=True, color="FFFFFF")
        cel.fill = PatternFill("solid", fgColor="1351B4")
        cel.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    larguras = [len(_rotulo_campo(k)) for k in chaves]
    for linha in linhas:
        ws.append([linha.get(k) for k in chaves])
        for i, k in enumerate(chaves):
            texto = "" if linha.get(k) is None else str(linha[k])
            larguras[i] = min(60, max(larguras[i], len(texto)))
    for i, larg in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(i)].width = larg + 2
    for row in ws.iter_rows(min_row=2):
        for cel in row:
            if isinstance(cel.value, float):
                cel.number_format = "#,##0.00"
    wb.save(caminho)


# ── consultas ───────────────────────────────────────────────────────────────

def dados_contratacoes(db, ano=None, modalidade=None, orgao=None):
    # relatório oficial: só o município do usuário (ver referencia=0 no
    # esquema — município de referência existe apenas para preços)
    where, args = ["referencia=0"], []
    if ano:
        where.append("ano=?")
        args.append(ano)
    if modalidade:
        where.append("modalidade_id=?")
        args.append(modalidade)
    if orgao:
        where.append("orgao_cnpj=?")
        args.append(orgao)
    sql_where = (" WHERE " + " AND ".join(where)) if where else ""
    linhas = [dict(r) for r in db.execute(
        f"""SELECT sequencial, ano, modalidade_nome,
                   json_extract(raw, '$.amparoLegal.nome') amparo,
                   objeto, unidade, valor_estimado, valor_homologado,
                   data_publicacao
            FROM contratacoes{sql_where}
            ORDER BY data_publicacao""", args)]
    tot_est = sum(l["valor_estimado"] or 0 for l in linhas)
    tot_hom = sum(l["valor_homologado"] or 0 for l in linhas)
    # deságio calculado só sobre processos com os dois valores
    pares = [(l["valor_estimado"], l["valor_homologado"]) for l in linhas
             if l["valor_estimado"] and l["valor_homologado"]]
    desagio = (1 - sum(h for _, h in pares) / sum(e for e, _ in pares)) * 100 \
        if pares else None
    return {"linhas": linhas,
            "totais": {"n": len(linhas), "estimado": tot_est,
                       "homologado": tot_hom, "desagio": desagio}}


def dados_contratos(db, ano=None, vigentes=False, orgao=None):
    where, args = [], []
    if vigentes:
        where.append("date(vigencia_fim) >= date('now','localtime')")
    elif ano:
        where.append("substr(data_publicacao,1,4)=?")
        args.append(str(ano))
    if orgao:
        where.append("orgao_cnpj=?")
        args.append(orgao)
    sql_where = (" WHERE " + " AND ".join(where)) if where else ""
    linhas = [dict(r) for r in db.execute(
        f"""SELECT numero_controle,
                   json_extract(raw, '$.numeroContratoEmpenho') numero,
                   json_extract(raw, '$.anoContrato') ano_contrato,
                   fornecedor_ni, fornecedor_nome, objeto, valor_global,
                   vigencia_inicio, vigencia_fim, data_publicacao
            FROM contratos{sql_where}
            ORDER BY data_publicacao""", args)]
    return {"linhas": linhas,
            "totais": {"n": len(linhas),
                       "valor": sum(l["valor_global"] or 0 for l in linhas)}}


def dados_atas(db, ano=None, vigentes=False, orgao=None):
    where, args = [], []
    if vigentes:
        where.append("date(vigencia_fim) >= date('now','localtime')")
    elif ano:
        where.append("substr(vigencia_inicio,1,4)=?")
        args.append(str(ano))
    if orgao:
        where.append("orgao_cnpj=?")
        args.append(orgao)
    sql_where = (" WHERE " + " AND ".join(where)) if where else ""
    linhas = [dict(r) for r in db.execute(
        f"""SELECT numero_controle,
                   json_extract(raw, '$.numeroAtaRegistroPreco') numero,
                   json_extract(raw, '$.anoAta') ano_ata,
                   json_extract(raw, '$.objetoContratacao') objeto,
                   contratacao_controle, vigencia_inicio, vigencia_fim,
                   json_extract(raw, '$.dataPublicacaoPncp') data_publicacao
            FROM atas{sql_where}
            ORDER BY vigencia_inicio""", args)]
    return {"linhas": linhas, "totais": {"n": len(linhas)}}


def dados_executivo(db, ano, orgao=None):
    ano = int(ano)
    og = " AND orgao_cnpj=?" if orgao else ""
    og_args = [orgao] if orgao else []
    modalidades = [dict(r) for r in db.execute(
        f"""SELECT modalidade_nome, COUNT(*) n,
                  SUM(valor_estimado) estimado, SUM(valor_homologado) homologado
           FROM contratacoes WHERE referencia=0 AND ano=?{og} GROUP BY 1
           ORDER BY COALESCE(SUM(COALESCE(valor_homologado, valor_estimado)),0)
           DESC""", [ano] + og_args)]
    meses = {r[0]: {"n": r[1], "valor": r[2] or 0} for r in db.execute(
        f"""SELECT substr(data_publicacao,6,2), COUNT(*),
                  SUM(COALESCE(valor_homologado, valor_estimado))
           FROM contratacoes
           WHERE referencia=0 AND ano=? AND data_publicacao IS NOT NULL{og}
           GROUP BY 1""", [ano] + og_args)}
    fornecedores = [dict(r) for r in db.execute(
        f"""SELECT fornecedor_nome, fornecedor_ni, COUNT(*) n,
                  SUM(COALESCE(valor_global,0)) total
           FROM contratos WHERE substr(data_publicacao,1,4)=?{og}
           GROUP BY fornecedor_ni ORDER BY total DESC LIMIT 10""",
        [str(ano)] + og_args)]
    vencendo = [dict(r) for r in db.execute(
        f"""SELECT 'Contrato' tipo, fornecedor_nome nome, objeto, vigencia_fim,
                  CAST(julianday(vigencia_fim) - julianday('now','localtime')
                       AS INTEGER) dias
           FROM contratos
           WHERE date(vigencia_fim) BETWEEN date('now','localtime')
                 AND date('now','localtime','+90 day'){og}
           UNION ALL
           SELECT 'Ata', json_extract(raw,'$.numeroAtaRegistroPreco') || '/' ||
                  json_extract(raw,'$.anoAta'),
                  json_extract(raw,'$.objetoContratacao'), vigencia_fim,
                  CAST(julianday(vigencia_fim) - julianday('now','localtime')
                       AS INTEGER)
           FROM atas
           WHERE date(vigencia_fim) BETWEEN date('now','localtime')
                 AND date('now','localtime','+90 day'){og}
           ORDER BY vigencia_fim""", og_args + og_args)]
    cards = dados_contratacoes(db, ano, orgao=orgao)["totais"]
    cards["contratos_vigentes"] = db.execute(
        f"SELECT COUNT(*) FROM contratos WHERE date(vigencia_fim)"
        f">=date('now','localtime'){og}", og_args).fetchone()[0]
    cards["atas_vigentes"] = db.execute(
        f"SELECT COUNT(*) FROM atas WHERE date(vigencia_fim)"
        f">=date('now','localtime'){og}", og_args).fetchone()[0]
    return {"ano": ano, "cards": cards, "modalidades": modalidades,
            "meses": meses, "fornecedores": fornecedores, "vencendo": vencendo}


def dados_fracionamento(db, ano, orgao=None, limites=None, janela=None):
    """Dispensas somadas por objeto de mesma natureza, contra o teto legal
    de cada uma (ver `teto_da_dispensa`).

    O agrupamento agora é por SIMILARIDADE do objeto (`_agrupar_por_
    similaridade`, motor portado do SGCD), que é o critério legal — "objeto
    de mesma natureza", art. 75, §1º — em vez do campo `unidade` do PNCP
    (que costuma trazer só o nome do órgão) ou de um radical fixo de
    palavras. Agrupa também por ÓRGÃO: Prefeitura e Câmara que dispensam a
    mesma coisa têm cada uma o seu limite, e somar as duas contra um teto só
    acusaria fracionamento (crime, art. 337-E do CP) onde há duas compras
    legais — e por TETO, porque o mesmo órgão pode ter dispensa de compra e
    de obra, com limites diferentes. Achado 2026-08-12, portado da varredura
    do licitarium-relatorios.

    `janela`: "exercicio" (padrão — ano civil de `ano`) ou "12"/"18"/"24"
    (meses corridos até hoje, período móvel — pega dispensa dividida na
    virada dez/jan, que o corte por exercício não vê; ver
    `JANELAS_FRAC_MESES`). Configurável em Configurações → Limites de
    dispensa. No modo móvel, `ano` só serve de rótulo — a seleção real é a
    janela de meses.

    Dispensa sem teto por valor (amparo em outro inciso do art. 75 ou em
    lei própria) não soma no termômetro nem no total — sai declarada à
    parte em `fora_do_limite_legal`, pra não desaparecer do relatório.
    """
    ano = int(ano)
    limites = limites or {}

    def _limite(valor, padrao):
        try:
            return float(valor)
        except (TypeError, ValueError):
            return padrao
    limite_compras = _limite(limites.get("compras"), LIMITE_PADRAO_COMPRAS)
    limite_obras = _limite(limites.get("obras"), LIMITE_PADRAO_OBRAS)
    og = " AND orgao_cnpj=?" if orgao else ""
    og_args = [orgao] if orgao else []
    meses = JANELAS_FRAC_MESES.get(str(janela))
    if meses:
        cond_data, arg_data = "date(data_publicacao) >= date('now','localtime',?)", f"-{meses} months"
    else:
        cond_data, arg_data = "ano=?", ano
    linhas = [dict(r) for r in db.execute(
        f"""SELECT numero_controle, sequencial, ano, orgao_cnpj, orgao_nome,
                   COALESCE(unidade,'(sem unidade)') unidade, objeto,
                   COALESCE(valor_homologado, valor_estimado, 0) valor,
                   data_publicacao,
                   json_extract(raw, '$.amparoLegal.nome') amparo
            FROM contratacoes
            WHERE referencia=0 AND {cond_data} AND modalidade_id=8{og}
            ORDER BY unidade, data_publicacao""", [arg_data] + og_args)]

    dispensas, fora, candidatas = [], [], []
    for l in linhas:
        inciso = _inciso_dispensa(l["amparo"])
        if inciso is None:
            fora.append(l)
            continue
        l["_teto"] = limite_compras if inciso == "II" else limite_obras
        l["_tipo"] = "compras" if inciso == "II" else "obras"
        dispensas.append(l)
        candidatas.append(l)
    unidades = sorted(
        (_linha_do_grupo(g) for g in _agrupar_por_similaridade(candidatas)),
        key=lambda u: -u["total"])

    return {"ano": ano, "unidades": unidades, "dispensas": dispensas,
            "fora_do_limite_legal": fora,
            "limite_compras": limite_compras, "limite_obras": limite_obras,
            "total": sum(d["valor"] or 0 for d in dispensas),
            "n": len(dispensas),
            "janela": str(janela) if meses else "exercicio"}


def dados_painel(db, ano, orgao=None, limites=None, janela=None):
    """Tudo o que o Painel mostra, numa consulta só por assunto.

    O painel tem três subabas — execução, análise e vigilância —, mas uma
    ida ao banco: a ponte JS custa mais que a consulta, e trocar de subaba
    não pode ir buscar dados de novo.

    `janela` segue `dados_fracionamento` — o card "Limite anual de
    dispensa" da Vigilância reusa o MESMO cálculo do relatório de
    Fracionamento (achado 2026-08-12/2026-08-25: existiam dois motores
    divergentes, um agrupando por unidade pro relatório e outro por objeto
    de 2 palavras pro Painel — unificado aqui, um cálculo só).
    """
    ano = int(ano)
    og = " AND orgao_cnpj=?" if orgao else ""
    og_args = [orgao] if orgao else []
    executivo = dados_executivo(db, ano, orgao)
    fracionamento = dados_fracionamento(db, ano, orgao, limites, janela)

    # ── execução: o ano corrente contra o anterior, no mesmo ponto do mês
    # Comparar o ano em curso com o ano anterior INTEIRO é aritmética do
    # calendário, não desempenho: em agosto, "caiu 67%" só diz que faltam
    # quatro meses. Quando o exercício pedido é o corrente, o anterior é
    # cortado no mesmo dia.
    og_c = " AND c.orgao_cnpj=?" if orgao else ""     # consultas com JOIN
    og_k = " AND k.orgao_cnpj=?" if orgao else ""
    hoje = date.today()
    parcial = ano == hoje.year
    corte = f"{ano - 1}-{hoje:%m-%d}" if parcial else f"{ano - 1}-12-31"
    ant = db.execute(
        f"""SELECT COUNT(*), SUM(valor_homologado), SUM(valor_estimado)
              FROM contratacoes
             WHERE referencia=0 AND ano=?
               AND (data_publicacao IS NULL OR substr(data_publicacao,1,10) <= ?)
               {og}""", [ano - 1, corte] + og_args).fetchone()
    anterior = {"n": ant[0], "homologado": ant[1] or 0, "estimado": ant[2] or 0}
    # homologado é homologado: o resumo executivo usa
    # COALESCE(homologado, estimado) para não zerar processo em andamento,
    # mas aqui as duas barras são comparadas lado a lado — misturar as duas
    # coisas numa delas faria o gráfico mentir sobre o que foi pago.
    mensal = {r[0]: r for r in db.execute(
        f"""SELECT CAST(substr(data_publicacao,6,2) AS INTEGER) mes,
                   COUNT(*) n, SUM(valor_estimado) est,
                   SUM(valor_homologado) hom
            FROM contratacoes
            WHERE referencia=0 AND ano=? AND data_publicacao IS NOT NULL{og}
            GROUP BY 1""", [ano] + og_args)}
    meses = [{"mes": m,
              "n": mensal[m][1] if m in mensal else 0,
              "estimado": (mensal[m][2] or 0) if m in mensal else 0,
              "valor": (mensal[m][3] or 0) if m in mensal else 0}
             for m in range(1, 13)]

    # ── análise: acumulado do ano e dos dois anteriores, mês a mês
    # a mesma consulta serve à série de homologado (análise) e à de
    # economia (economia): um SELECT a mais (valor_estimado) em vez de
    # duas idas ao banco por ano
    series, series_economia = {}, {}
    for a in (ano - 2, ano - 1, ano):
        # mesma regra do gráfico mensal: acumulado de homologado é só do
        # que foi efetivamente homologado
        por_mes = {r[0]: (r[1] or 0, r[2] or 0) for r in db.execute(
            f"""SELECT CAST(substr(data_publicacao,6,2) AS INTEGER),
                       SUM(valor_homologado), SUM(valor_estimado)
                FROM contratacoes
                WHERE referencia=0 AND ano=? AND data_publicacao IS NOT NULL{og}
                GROUP BY 1""", [a] + og_args)}
        acumulado, acumulado_economia = [], []
        total, total_economia = 0, 0
        for m in range(1, 13):
            hom, est = por_mes.get(m, (0, 0))
            total += hom
            acumulado.append(total)
            total_economia += est - hom
            acumulado_economia.append(total_economia)
        series[a] = acumulado
        series_economia[a] = acumulado_economia

    # deságio por modalidade: quanto o certame economizou sobre o estimado
    desagios = []
    for r in db.execute(
            f"""SELECT modalidade_nome, COUNT(*) n,
                       SUM(valor_estimado) est, SUM(valor_homologado) hom
                FROM contratacoes
                WHERE referencia=0 AND ano=? AND valor_estimado > 0
                  AND valor_homologado IS NOT NULL{og}
                GROUP BY 1""", [ano] + og_args):
        desagios.append({"modalidade": r[0], "n": r[1],
                         "estimado": r[2] or 0, "homologado": r[3] or 0,
                         "economizado": (r[2] or 0) - (r[3] or 0),
                         "pct": (1 - (r[3] or 0) / r[2]) * 100 if r[2] else 0})
    # ordena pelo que o gráfico desenha (economizado), não pelo estimado —
    # achado 2026-08-12, portado do licitarium-relatorios: as outras três
    # listas (família/categoria/fornecedor) já ordenam por economizado, só
    # esta estava com a métrica errada
    desagios.sort(key=lambda d: -d["economizado"])

    # economia por família de item e por categoria bruta do PNCP — a mesma
    # pergunta do deságio por modalidade, só que descendo ao item; uma
    # consulta só em `itens` alimenta as duas listas
    itens_economia = [dict(r) for r in db.execute(
        f"""SELECT descricao, categoria, material_servico,
                   fornecedor_ni, fornecedor_nome,
                   valor_total_estimado est, valor_total_homologado hom
              FROM itens
             WHERE referencia=0 AND ano=? AND valor_total_estimado > 0
               AND valor_total_homologado IS NOT NULL{og}""",
        [ano] + og_args)]
    por_familia, por_categoria, por_fornecedor = {}, {}, {}
    for it in itens_economia:
        chave = pca_builder.chave_agrupamento(it["descricao"], palavras=2) \
            or "(sem descrição)"
        alvo = por_familia.setdefault(
            chave, {"nome": chave, "n": 0, "estimado": 0.0, "homologado": 0.0})
        alvo["n"] += 1
        alvo["estimado"] += it["est"] or 0
        alvo["homologado"] += it["hom"] or 0
        cat = _categoria_util(it["categoria"], it["material_servico"])
        alvo_cat = por_categoria.setdefault(
            cat, {"nome": cat, "n": 0, "estimado": 0.0, "homologado": 0.0})
        alvo_cat["n"] += 1
        alvo_cat["estimado"] += it["est"] or 0
        alvo_cat["homologado"] += it["hom"] or 0
        # agrupa pelo documento, não pelo nome: a mesma empresa aparece com
        # grafias diferentes entre processos (o `ni` é CNPJ ou CPF — ver
        # `documento()`). Item sem fornecedor não é atribuível a ninguém.
        if it["fornecedor_ni"]:
            alvo_forn = por_fornecedor.setdefault(
                it["fornecedor_ni"],
                {"nome": it["fornecedor_nome"] or it["fornecedor_ni"],
                 "ni": it["fornecedor_ni"], "n": 0,
                 "estimado": 0.0, "homologado": 0.0})
            alvo_forn["n"] += 1
            alvo_forn["estimado"] += it["est"] or 0
            alvo_forn["homologado"] += it["hom"] or 0
    for grupo in (por_familia, por_categoria, por_fornecedor):
        for v in grupo.values():
            v["economizado"] = v["estimado"] - v["homologado"]
            v["pct"] = (1 - v["homologado"] / v["estimado"]) * 100 \
                if v["estimado"] else 0
    por_familia = sorted(por_familia.values(), key=lambda o: -o["economizado"])
    por_categoria = sorted(por_categoria.values(), key=lambda o: -o["economizado"])
    por_fornecedor = sorted(por_fornecedor.values(),
                            key=lambda o: -o["economizado"])

    # concentração: quanto do valor está nos maiores fornecedores
    valores = [r[0] or 0 for r in db.execute(
        f"""SELECT SUM(COALESCE(valor_global,0)) t FROM contratos
            WHERE substr(data_publicacao,1,4)=?{og}
            GROUP BY fornecedor_ni ORDER BY t DESC""",
        [str(ano)] + og_args)]
    total_contratado = sum(valores)
    curva, acumulado = [], 0
    for v in valores:
        acumulado += v
        curva.append(acumulado / total_contratado * 100 if total_contratado else 0)

    # calor: processos por mês e modalidade, com a cauda somada em "Outras"
    principais = [m["modalidade_nome"] for m in executivo["modalidades"][:3]]
    calor = {nome: [0] * 12 for nome in principais + ["Outras"]}
    for r in db.execute(
            f"""SELECT modalidade_nome, CAST(substr(data_publicacao,6,2) AS INTEGER),
                       COUNT(*)
                FROM contratacoes
                WHERE referencia=0 AND ano=? AND data_publicacao IS NOT NULL{og}
                GROUP BY 1,2""", [ano] + og_args):
        linha = calor[r[0]] if r[0] in calor else calor["Outras"]
        if r[1] and 1 <= r[1] <= 12:
            linha[r[1] - 1] += r[2]

    # ── vigilância: o que exige ação
    funil = {
        "publicadas": executivo["cards"]["n"],
        # `contratacoes` e `itens` têm as duas uma coluna orgao_cnpj: sem o
        # prefixo, filtrar por órgão fazia o SQLite recusar a consulta
        # inteira ("ambiguous column name") e o painel não abria
        "com_resultado": db.execute(
            f"""SELECT COUNT(DISTINCT c.numero_controle) FROM contratacoes c
                 JOIN itens i ON i.contratacao_controle = c.numero_controle
                WHERE c.referencia=0 AND c.ano=?
                  AND i.valor_unitario_homologado IS NOT NULL{og_c}""",
            [ano] + og_args).fetchone()[0],
        "com_contrato": db.execute(
            f"""SELECT COUNT(DISTINCT contratacao_controle) FROM contratos k
                WHERE k.contratacao_controle IN (
                  SELECT numero_controle FROM contratacoes
                   WHERE referencia=0 AND ano=?){og_k}""",
            [ano] + og_args).fetchone()[0],
        # vigentes DO EXERCÍCIO: contar todos os contratos vigentes, de
        # qualquer ano, fazia a última etapa do funil ficar maior que a
        # primeira — as quatro barras precisam falar do mesmo conjunto
        "vigentes": db.execute(
            f"""SELECT COUNT(*) FROM contratos k
                 WHERE date(k.vigencia_fim) >= date('now','localtime')
                   AND k.contratacao_controle IN (
                     SELECT numero_controle FROM contratacoes
                      WHERE referencia=0 AND ano=?){og_k}""",
            [ano] + og_args).fetchone()[0],
    }
    # processo publicado há muito tempo e sem resultado é pendência, não
    # estatística: costuma ser homologação que o órgão esqueceu de publicar
    paradas = db.execute(
        f"""SELECT COUNT(*) FROM contratacoes c
             WHERE c.referencia=0 AND c.valor_homologado IS NULL
               AND date(c.data_publicacao) < date('now','localtime','-90 day')
               AND c.ano=?{og_c}""", [ano] + og_args).fetchone()[0]
    propostas = db.execute(
        f"""SELECT COUNT(*) FROM contratacoes
             WHERE referencia=0
               AND datetime(data_encerramento_proposta)
                   >= datetime('now','localtime'){og}""",
        og_args).fetchone()[0]
    # o card "Limite anual" reusa o mesmo agrupamento por similaridade do
    # relatório de Fracionamento (`fracionamento`, já calculado acima) —
    # não recalcula nada aqui (achado 2026-08-25: eram dois motores
    # divergentes, um por objeto de Nº fixo de palavras, outro por unidade)
    objetos = fracionamento["unidades"]
    perto_do_limite = [o for o in objetos if o["pct"] >= 75]

    return {
        "ano": ano,
        "comparacao_parcial": parcial,
        "alertas": {"perto_do_limite": len(perto_do_limite),
                    # o clique no chip filtra a lista por estes PROCESSOS —
                    # não por "toda dispensa do ano". Desde que o
                    # agrupamento virou similaridade (2026-08-25), o grupo
                    # não é mais recalculável em SQL por um radical fixo;
                    # a lista de numero_controle vai explícita
                    "objetos_perto_do_limite": [nc for o in perto_do_limite
                                                for nc in o["numeros_controle"]],
                    "acima_do_limite": sum(1 for o in objetos
                                           if o["pct"] > 100),
                    # contrato e ata vivem em telas diferentes — um alerta só
                    # não dá pra clicar e ir aos dois ao mesmo tempo
                    "vencendo_contratos": sum(
                        1 for v in executivo["vencendo"]
                        if v["tipo"] == "Contrato" and (v["dias"] or 0) <= 60),
                    "vencendo_atas": sum(
                        1 for v in executivo["vencendo"]
                        if v["tipo"] == "Ata" and (v["dias"] or 0) <= 60),
                    "propostas": propostas, "paradas": paradas},
        "execucao": {"cards": executivo["cards"], "meses": meses,
                     "modalidades": executivo["modalidades"],
                     "fornecedores": executivo["fornecedores"],
                     "vencendo": executivo["vencendo"],
                     "homologado_anterior": anterior["homologado"],
                     "n_anterior": anterior["n"]},
        "analise": {"series": {str(a): v for a, v in series.items()},
                    "desagios": desagios, "curva": curva,
                    "fornecedores_total": len(valores),
                    "calor": calor, "meses_calor": list(range(1, 13))},
        "vigilancia": {"funil": funil, "limites": objetos[:6],
                       "limite_compras": fracionamento["limite_compras"],
                       "agenda": executivo["vencendo"][:40]},
        "economia": {
            "estimado": executivo["cards"]["estimado"],
            "homologado": executivo["cards"]["homologado"],
            "economizado": (executivo["cards"]["estimado"] or 0)
                - (executivo["cards"]["homologado"] or 0),
            "pct": executivo["cards"]["desagio"],
            "estimado_anterior": anterior["estimado"],
            "homologado_anterior": anterior["homologado"],
            "economizado_anterior": (anterior["estimado"] - anterior["homologado"])
                if anterior["estimado"] and anterior["homologado"] else None,
            "por_modalidade": desagios,
            "por_familia": por_familia[:10],
            "por_categoria": por_categoria[:10],
            "por_fornecedor": por_fornecedor[:10],
            "series": {str(a): v for a, v in series_economia.items()},
        },
    }


def _blocos(ids, tamanho=400):
    """Fatia ids para caber no limite de parâmetros do SQLite."""
    ids = [str(i) for i in (ids or []) if i]
    return [ids[i:i + tamanho] for i in range(0, len(ids), tamanho)]


def mes_por_extenso(competencia_):
    """"2026-06" vira "jun/2026", que é como o documento fala.

    O ano é validado como número junto com o mês: antes só o mês passava
    por `int()`, e um ano com marcação HTML saía inteiro na prosa do
    relatório — que é aberto no navegador real (auditoria de segurança,
    2026-08-09). Competência fora do formato não vira texto: some.
    """
    if not competencia_:
        return None
    partes = str(competencia_).split("-")
    if len(partes) != 2:
        return None
    try:
        ano, mes = int(partes[0]), int(partes[1])
    except ValueError:
        return None
    if not 1 <= mes <= 12:
        return None
    return f"{MESES_NOME[mes - 1]}/{ano}"


# ── render ──────────────────────────────────────────────────────────────────

# O documento impresso NÃO segue o tema da tela (mudança consciente da
# v1.20.0; a v1.14.4 tinha feito o contrário, ver CHANGELOG). Papel que vai
# ao Tribunal de Contas é peça institucional do município, não vitrine da
# ferramenta: fundo branco, grafite no lugar do vinho/dourado, réguas
# discretas. Os três temas continuam valendo integralmente na tela — e por
# isso não existe parâmetro de tema aqui: o documento não tem como seguir a
# tela nem por engano.
PALETA_DOCUMENTO = dict(
    bg="#ffffff", superficie="#ffffff", zebra="#f6f7f8",
    cabecalho="#eef0f2", texto="#17181a", suave="#5b6066",
    borda="#d3d6da", acento="#1f2933", detalhe="#b8bec4",
    alerta="#a6231b", atencao="#7a5c0e", azul="#2f4b7c", verde="#2c6149")

_VARS = " ".join(f"--{chave}:{cor};" for chave, cor in PALETA_DOCUMENTO.items())

# Selo de procedência (achado 2026-08-13, portado do diagnóstico de
# identidade do licitarium-relatorios): a cor do documento codifica o TIPO
# de trabalho — como a CGU distingue apuração/avaliação/consultoria na
# capa — não decora, informa antes de abrir. Cadastral (relação factual) em
# preto/grafite; analítico (execução, economia) em azul; vigilância (teto,
# fracionamento) reaproveita o alerta — é literalmente o assunto do
# documento; planejamento (PCA, preços) em verde.
CATEGORIA_RELATORIO = {
    "contratacoes": ("Cadastral", "acento"),
    "contratos": ("Cadastral", "acento"),
    "atas": ("Cadastral", "acento"),
    "executivo": ("Analítico", "azul"),
    "economia": ("Analítico", "azul"),
    "fracionamento": ("Vigilância", "alerta"),
    "minuta_pca": ("Planejamento", "verde"),
}


def _acervo_atual(db):
    """Data/hora do dado mais recente sincronizado — a "prova de
    procedência" que entra na faixa do cabeçalho e no rodapé de cada
    relatório: de que fotografia do acervo este documento saiu.
    """
    r = db.execute(
        """SELECT MAX(v) FROM (
             SELECT MAX(sync_em) v FROM contratacoes
             UNION SELECT MAX(sync_em) FROM contratos
             UNION SELECT MAX(sync_em) FROM atas)""").fetchone()
    return r[0] if r else None


def _css(paisagem, papel="A4"):
    return f"""
  :root {{ {_VARS} }}
  @page {{
    size: {papel} {"landscape" if paisagem else "portrait"}; margin: 1.6cm 1.4cm;
    @top-center {{ content: string(titulo); font-size: 8pt; color: #6f5b3e; }}
    @bottom-right {{ content: "Página " counter(page) " de " counter(pages);
                     font-size: 8pt; color: #6f5b3e; }}
  }}
  * {{ margin:0; padding:0; box-sizing:border-box; }}
  body {{ font-family:'Segoe UI',system-ui,sans-serif; color:var(--texto);
          background:var(--bg); font-size:13px; line-height:1.45; }}
  /* Sem `max-width` em pixel na impressão: A4 paisagem com margem de 1,4 cm
     tem ~1017 px úteis, e o antigo 1080px transbordava a caixa do papel — os
     gráficos (largura 100% do bloco) saíam 63 px pela direita e eram
     cortados. Quem define a largura é a `@page` acima; aqui a página só
     ocupa o que sobrou. O `max-width` em px foi calibrado para pixel de
     tela, e o defeito escondeu-se porque o teste antigo gerava o PDF
     passando `margin` no `page.pdf()`, ignorando a `@page` do CSS. */
  .pagina {{ max-width:100%; margin:0 auto; padding:26px 30px 50px; }}
  header {{ display:flex; align-items:center; gap:18px; padding-bottom:14px;
            border-bottom:1px solid var(--borda); margin-bottom:16px; }}
  h1 {{ font-family:Georgia,serif; font-size:21px; font-weight:400;
        string-set: titulo content(); }}
  .meta {{ font-size:11.5px; color:var(--suave); margin-top:3px; }}
  h2 {{ font-family:Georgia,serif; font-size:15px; font-weight:400;
        color:var(--acento); margin:20px 0 8px; break-after:avoid; }}
  table {{ border-collapse:collapse; width:100%; font-size:11.5px; }}
  th, td {{ border:1px solid var(--borda); padding:5px 8px; text-align:left;
            vertical-align:middle; }}
  th {{ background:var(--cabecalho); font-size:10px; letter-spacing:.05em;
        text-transform:uppercase; }}
  tr {{ break-inside:avoid; }}
  tbody tr:nth-child(even) td {{ background:var(--zebra); }}
  /* linha fora da faixa esperada (Pesquisa de Preços): cor nunca sozinha
     (WCAG 1.4.1) — o "*" no valor e a nota de rodapé carregam o mesmo
     alerta pra quem imprime em P&B ou não distingue vermelho */
  tr.fora td {{ color:var(--alerta); }}
  /* colunas curtas (valores, datas, qtde): centro nos dois eixos */
  td.num, th.num {{ text-align:center; font-variant-numeric:tabular-nums;
                    white-space:nowrap; }}
  /* centro com quebra de linha permitida (textos curtos não-numéricos) */
  td.ctr, th.ctr {{ text-align:center; }}
  tfoot td {{ background:var(--cabecalho); font-weight:600; }}
  .obj {{ text-transform:uppercase; text-align:justify; hyphens:auto; }}
  /* nome de fornecedor quebra feio; a coluna cede espaço da descrição */
  td.forn, th.forn {{ text-align:center; min-width:170px; }}
  /* município e unidade em uma linha só: "Paulo de Faria" e "Fardo 64,00 RO"
     quebravam no meio, e a descrição tem folga para ceder */
  td.muni, th.muni {{ text-align:center; white-space:nowrap; }}
  td.unid, th.unid {{ text-align:center; white-space:nowrap; }}
  /* link para a página oficial: discreto no papel, clicável no PDF */
  td.proc a {{ color:var(--acento); text-decoration:none;
               border-bottom:1px dotted var(--acento); }}
  .cards {{ display:flex; gap:10px; margin-bottom:6px; }}
  /* dispersão da série: leitura de apoio aos números em destaque */
  p.disp {{ margin:0 0 6px; font-size:10.5px; color:var(--suave);
            break-inside:avoid; }}
  p.disp b {{ color:var(--texto); }}
  td.sem-motivo {{ color:var(--alerta); font-style:italic; }}
  .card {{ background:var(--superficie); border:1px solid var(--borda);
           border-top:3px solid var(--cor-categoria, var(--acento));
           border-radius:3px;
           padding:9px 12px 10px; break-inside:avoid; flex:1 1 auto; }}
  .card .n {{ font-family:Georgia,serif; font-size:17px; color:var(--acento);
              white-space:nowrap; }}
  /* versalete no lugar do negrito (achado 2026-08-13, portado do
     diagnóstico de identidade): o texto de origem já vem em minúsculas —
     small-caps lê como rótulo sem gritar tanto quanto caixa alta */
  .card .l {{ font-size:10.5px; letter-spacing:.03em; font-variant:small-caps;
              color:var(--suave); margin-top:2px; }}
  /* selo de procedência: etiqueta da categoria acima do cabeçalho, régua e
     faixa de acervo abaixo dele — mesmo padrão CGU (cor = tipo de
     trabalho), sem competir com o brasão do município */
  .etiqueta {{ display:inline-block; font-size:9px;
               font-family:'Segoe UI',system-ui,sans-serif;
               letter-spacing:.12em; text-transform:uppercase; color:#fff;
               background:var(--cor-categoria, var(--acento));
               padding:3px 9px; border-radius:2px; margin-bottom:10px; }}
  .regua {{ height:2px; border:none;
            background:var(--cor-categoria, var(--acento)); margin:14px 0 6px; }}
  .faixa-acervo {{ font-family:Consolas,monospace; font-size:9.5px;
                    letter-spacing:.03em; color:var(--suave);
                    margin-bottom:16px; }}
  .barra {{ background:var(--detalhe); height:10px; display:inline-block;
            vertical-align:middle; border-radius:2px; }}
  .caixa-aviso {{ background:var(--superficie); border:1px solid var(--borda);
                  border-left:4px solid var(--alerta); border-radius:3px;
                  padding:10px 14px; font-size:11.5px; margin-bottom:12px;
                  break-inside:avoid; }}
  .farol-alerta {{ color:var(--alerta); font-weight:600; }}
  .farol-atencao {{ color:var(--atencao); font-weight:600; }}
  footer {{ margin-top:22px; padding-top:10px;
            border-top:1px solid var(--borda);
            font-size:10.5px; color:var(--suave); display:flex;
            justify-content:space-between; }}
  .no-print {{ position:fixed; top:14px; right:14px; }}
  .no-print button {{ font-size:14px; padding:8px 14px; cursor:pointer;
    background:var(--acento); color:var(--superficie); border:none;
    border-radius:3px; }}
  @media print {{ body {{ background:var(--bg); font-size:10pt; }}
    tbody tr:nth-child(even) td {{ background:var(--zebra); }}
    .pagina {{ max-width:none; padding:0; }} .no-print {{ display:none; }} }}
"""


def _pagina(titulo_doc, corpo, municipio, uf, periodo_txt, paisagem,
            papel="A4", estilo_extra="", brasao=None, categoria=None,
            acervo=None):
    agora = datetime.now().strftime("%d/%m/%Y %H:%M")
    # o brasão do município (Configurações) toma o lugar do estandarte do
    # Licitarium na frente do documento — o produto continua assinado no
    # rodapé ("LICITARIVM · SVB HASTA PVBLICA"), só quem imprime muda
    marca = (f'<img src="{_e(brasao)}" alt="Brasão do município"'
             f' style="height:88px;width:auto">' if brasao else ESTANDARTE)
    # selo de procedência (achado 2026-08-13): sem `categoria`, nenhuma
    # etiqueta/régua/faixa aparece — o Painel e a ficha de detalhe (que já
    # têm identidade visual própria) chamam `_pagina` sem esse parâmetro e
    # continuam exatamente como antes
    estilo_body = (f' style="--cor-categoria:var(--{categoria[1]});"'
                   if categoria else '')
    etiqueta_html = f'<span class="etiqueta">{_e(categoria[0])}</span>' \
        if categoria else ''
    faixa_html = ''
    if categoria and acervo:
        hash_curto = hashlib.sha1(str(acervo).encode()).hexdigest()[:6].upper()
        faixa_html = (f'<hr class="regua"><div class="faixa-acervo">'
                      f'Acervo sincronizado em {data_br(acervo)} · {hash_curto}'
                      f'</div>')
    elif categoria:
        faixa_html = '<hr class="regua">'
    rodape_proc = (f'Apurado a partir do PNCP · acervo sincronizado em '
                   f'{data_br(acervo)}' if acervo else
                   'Documento gerado automaticamente a partir de dados '
                   'públicos do PNCP')
    return f"""<!DOCTYPE html>
<html lang="pt-BR"><head><meta charset="utf-8">
<title>{_e(titulo_doc)}</title>
<style>{_css(paisagem, papel)}{estilo_extra}</style></head>
<body{estilo_body}>
<div class="no-print"><button onclick="print()">🖨 Imprimir</button></div>
<div class="pagina">
{etiqueta_html}
<header>{marca}
  <div><h1>{_e(titulo_doc)}</h1>
  <div class="meta">{_e(municipio)} — {_e(uf)} · {_e(periodo_txt)}<br>
  Fonte: Portal Nacional de Contratações Públicas (PNCP) · Lei 14.133/2021<br>
  Gerado pelo Licitarium em {agora}</div></div>
</header>
{faixa_html}
{corpo}
<footer><span>LICITARIVM · SVB HASTA PVBLICA</span>
<span>{rodape_proc}</span></footer>
</div></body></html>"""


def render_contratacoes(d, municipio, uf, periodo_txt, brasao=None,
                        categoria=None, acervo=None):
    linhas = "".join(f"""<tr>
      <td class="ctr">{_e(l['sequencial'])}/{_e(l['ano'])}</td>
      <td class="ctr">{_e(l['modalidade_nome'])}</td>
      <td class="ctr">{_e(l['amparo'])}</td>
      <td class="obj">{_e(l['objeto'])}</td>
      <td class="ctr">{_e(l['unidade'])}</td>
      <td class="num">{moeda(l['valor_estimado'])}</td>
      <td class="num">{moeda(l['valor_homologado'])}</td>
      <td class="num">{data_br(l['data_publicacao'])}</td></tr>"""
      for l in d["linhas"])
    t = d["totais"]
    desagio = f" · Deságio médio: {t['desagio']:.1f}%".replace(".", ",") \
        if t["desagio"] is not None else ""
    corpo = f"""<div class="caixa-aviso">Relação de <b>todas as contratações
publicadas</b> pelo município no período, extraída do PNCP — inclui
processos em andamento (sem homologação) e já concluídos. Valor estimado é
o do edital; valor homologado é o valor final, quando já há resultado.</div>
<table>
<thead><tr><th class="ctr">Processo</th><th class="ctr">Modalidade</th>
<th class="ctr">Amparo legal</th>
<th>Objeto</th><th class="ctr">Unidade</th><th class="num">Valor estimado</th>
<th class="num">Valor homologado</th><th class="num">Publicação</th></tr></thead>
<tbody>{linhas or '<tr><td colspan="8">Nenhum registro no período.</td></tr>'}</tbody>
<tfoot><tr><td colspan="5">Total: {t['n']} contratações{desagio}</td>
<td class="num">{moeda(t['estimado'])}</td>
<td class="num">{moeda(t['homologado'])}</td><td></td></tr></tfoot></table>"""
    titulo = f"{TITULOS['contratacoes']} — {municipio} — {periodo_txt}"
    return _pagina(titulo, corpo, municipio, uf, periodo_txt, paisagem=True,
                   brasao=brasao, categoria=categoria, acervo=acervo)


def render_contratos(d, municipio, uf, periodo_txt, brasao=None,
                     categoria=None, acervo=None):
    linhas = "".join(f"""<tr>
      <td class="ctr">{_e(num_contrato(l['numero'], l['ano_contrato'])
                          or l['numero_controle'])}</td>
      <td class="ctr">{_e(l['fornecedor_nome'])}<br>
          <small>{_e(documento(l['fornecedor_ni']))}</small></td>
      <td class="obj">{_e(l['objeto'])}</td>
      <td class="num">{moeda(l['valor_global'])}</td>
      <td class="num">{data_br(l['vigencia_inicio'])} – {data_br(l['vigencia_fim'])}</td>
      <td class="num">{data_br(l['data_publicacao'])}</td></tr>"""
      for l in d["linhas"])
    t = d["totais"]
    corpo = f"""<div class="caixa-aviso">Relação de <b>contratos firmados</b>
pelo município, extraída do PNCP. Contrato é o instrumento definitivo —
decorrente de uma contratação já homologada, com fornecedor e vigência
definidos.</div>
<table>
<thead><tr><th class="ctr">Contrato</th><th class="ctr">Fornecedor</th><th>Objeto</th>
<th class="num">Valor global</th><th class="num">Vigência</th>
<th class="num">Publicação</th></tr></thead>
<tbody>{linhas or '<tr><td colspan="6">Nenhum registro no período.</td></tr>'}</tbody>
<tfoot><tr><td colspan="3">Total: {t['n']} contratos</td>
<td class="num">{moeda(t['valor'])}</td><td colspan="2"></td></tr></tfoot></table>"""
    titulo = f"{TITULOS['contratos']} — {municipio} — {periodo_txt}"
    return _pagina(titulo, corpo, municipio, uf, periodo_txt, paisagem=True,
                   brasao=brasao, categoria=categoria, acervo=acervo)


def render_atas(d, municipio, uf, periodo_txt, brasao=None, categoria=None,
                acervo=None):
    linhas = "".join(f"""<tr>
      <td class="ctr">{_e(l['numero'])}/{_e(l['ano_ata'])}</td>
      <td class="ctr">{_e(l['contratacao_controle'])}</td>
      <td class="obj">{_e(l['objeto'])}</td>
      <td class="num">{data_br(l['vigencia_inicio'])} – {data_br(l['vigencia_fim'])}</td>
      <td class="num">{data_br(l['data_publicacao'])}</td></tr>"""
      for l in d["linhas"])
    corpo = f"""<div class="caixa-aviso">Relação de <b>atas de registro de
preços</b> vigentes ou já encerradas, extraída do PNCP. A ata registra o
preço para contratações futuras dentro do prazo de vigência — não é, em
si, uma despesa executada.</div>
<table>
<thead><tr><th class="ctr">Ata</th><th class="ctr">Contratação de origem</th><th>Objeto</th>
<th class="num">Vigência</th><th class="num">Publicação</th></tr></thead>
<tbody>{linhas or '<tr><td colspan="5">Nenhum registro no período.</td></tr>'}</tbody>
<tfoot><tr><td colspan="5">Total: {d['totais']['n']} atas</td></tr></tfoot></table>"""
    titulo = f"{TITULOS['atas']} — {municipio} — {periodo_txt}"
    return _pagina(titulo, corpo, municipio, uf, periodo_txt, paisagem=True,
                   brasao=brasao, categoria=categoria, acervo=acervo)


def render_fracionamento(d, municipio, uf, brasao=None, categoria=None,
                         acervo=None):
    def farol(pct):
        if pct >= 100:
            return '<span class="farol-alerta">ACIMA DO LIMITE</span>'
        if pct >= 75:
            return '<span class="farol-atencao">Atenção</span>'
        return "ok"
    # a coluna Órgão só aparece quando há mais de um com dispensa: sem ela,
    # duas linhas do mesmo objeto pareceriam duplicata, quando são dois
    # entes com teto próprio cada um
    varios_orgaos = len({u["orgao_nome"] for u in d["unidades"]}) > 1
    unid = "".join(f"""<tr>{
      f'<td>{_e(u["orgao_nome"])}</td>' if varios_orgaos else ''}
      <td>{_e(u['objeto'])}</td>
      <td>{'Obras' if u['tipo'] == 'obras' else 'Compras'}</td>
      <td class="num">{u['n']}</td>
      <td class="num">{moeda(u['total'])}</td>
      <td class="num">{u['pct']:.0f}%</td>
      <td class="ctr">{farol(u['pct'])}</td></tr>""" for u in d["unidades"])
    disp = "".join(f"""<tr><td class="ctr">{_e(l['sequencial'])}/{_e(l['ano'])}</td>
      <td class="ctr">{_e(l['unidade'])}</td>
      <td class="obj">{_e(l['objeto'])}</td>
      <td class="num">{moeda(l['valor'])}</td>
      <td class="num">{data_br(l['data_publicacao'])}</td></tr>"""
      for l in d["dispensas"])
    n_colunas = 6 if varios_orgaos else 5
    fora = d.get("fora_do_limite_legal") or []
    nota_fora = (f'<p class="nota">Cada linha é medida contra o teto do seu '
                 f'próprio inciso — obra tem limite dobrado em relação a '
                 f'compra, e dispensa do mesmo órgão nunca soma com a de '
                 f'outro. Outras {len(fora)} dispensa'
                 f'{"s" if len(fora) != 1 else ""} do período ficaram '
                 f'fora deste quadro por não terem teto por valor — amparo '
                 f'em outro inciso do art. 75 (emergência, licitação '
                 f'deserta, agricultura familiar etc.) ou em lei própria '
                 f'decorre da natureza do objeto, não do valor.</p>'
                 if fora else '')
    janela = d.get("janela") or "exercicio"
    janela_label = ("exercício financeiro — 1º de janeiro a 31 de dezembro"
                    if janela == "exercicio"
                    else f"período móvel — últimos {janela} meses corridos")
    corpo = f"""<div class="caixa-aviso">Instrumento de <b>autocontrole
interno</b>. A soma por objeto semelhante é um termômetro: o enquadramento
legal do fracionamento considera despesas de <b>mesma natureza</b> (art. 75,
§1º, Lei 14.133/2021), avaliação que cabe ao gestor — o agrupamento aqui é
por similaridade textual do objeto, não substitui essa análise. Janela de
análise: <b>{janela_label}</b>. Limites parametrizados nas configurações —
confira o decreto de atualização vigente.
Limite adotado para compras/serviços (art. 75, II): <b>{moeda(d['limite_compras'])}</b> ·
obras/serviços de engenharia (art. 75, I): <b>{moeda(d['limite_obras'])}</b>.
Cada dispensa é medida contra o teto do próprio órgão — nunca somada com a
de outro ente.</div>
<div class="cards">
<div class="card"><div class="n">{d['n']}</div><div class="l">dispensas no período</div></div>
<div class="card"><div class="n">{moeda(d['total'])}</div><div class="l">total em dispensas</div></div>
</div>
<h2>Soma de dispensas por objeto semelhante × teto legal</h2>
<div class="card">{_grafico_limites(d["unidades"])}</div>
<table><thead><tr>{
  '<th>Órgão</th>' if varios_orgaos else ''}<th>Objeto</th><th>Teto</th>
<th class="num">Dispensas</th>
<th class="num">Total</th><th class="num">% do teto</th>
<th class="ctr">Situação</th></tr></thead>
<tbody>{unid or f'<tr><td colspan="{n_colunas}">Nenhuma dispensa com teto por valor no período.</td></tr>'}</tbody></table>
{nota_fora}
<h2>Dispensas do período (para agrupamento por natureza pelo gestor)</h2>
<table><thead><tr><th class="ctr">Processo</th><th class="ctr">Unidade</th>
<th>Objeto</th><th class="num">Valor</th><th class="num">Publicação</th></tr></thead>
<tbody>{disp or '<tr><td colspan="5">Nenhuma dispensa no período.</td></tr>'}</tbody></table>"""
    titulo = f"{TITULOS['fracionamento']} {d['ano']} — {municipio}"
    return _pagina(titulo, corpo, municipio, uf,
                   f"Exercício {d['ano']} · uso interno", paisagem=True,
                   estilo_extra=CSS_PAINEL, brasao=brasao,
                   categoria=categoria, acervo=acervo)


def render_minuta_pca(d, municipio, uf, brasao=None, categoria=None,
                      acervo=None):
    linhas = "".join(f"""<tr>
      <td class="num">{i+1}</td>
      <td class="obj">{_e(l['descricao'])}</td>
      <td class="ctr">{_e(l['categoria'])}</td>
      <td class="ctr">{_e(l['unidade'])}</td>
      <td class="num">{l['quantidade'] or 0:.2f}</td>
      <td class="num">{moeda(l['valor_unitario'])}</td>
      <td class="num">{moeda(l['valor_total'])}</td>
      <td class="ctr">{f"<b>{l['abc']}</b>" if l['abc'] == 'A' else l['abc']}</td></tr>"""
      for i, l in enumerate(d["itens"]))
    p = d.get("parametros") or {}
    base = {"media": "média dos exercícios", "ultimo": "último exercício",
            "maior": "maior exercício", "soma": "soma do período"}.get(
                p.get("base"), p.get("base", "—"))
    est = {"mediana": "mediana", "media": "média", "recente": "mais recente",
           "menor": "menor"}.get(p.get("estatistica"), p.get("estatistica", "—"))
    # curva ABC (pca_builder.classificar_abc, já calculada em listar_minuta):
    # onde vale gastar o tempo de revisão — poucos itens costumam responder
    # pela maior parte do valor. Achado do usuário (2026-08-08): o cálculo já
    # existia e alimentava a tela de Montar PCA, mas não aparecia no documento
    abc = {}
    valor_itens = sum(l["valor_total"] for l in d["itens"]) or 1
    for l in d["itens"]:
        c = abc.setdefault(l["abc"], {"n": 0, "valor": 0.0})
        c["n"] += 1
        c["valor"] += l["valor_total"]
    linha_abc = " · ".join(
        f"{abc[c]['n']} {'item' if abc[c]['n'] == 1 else 'itens'} classe {c}"
        f" = {abc[c]['valor'] / valor_itens * 100:.0f}% do valor"
        for c in ("A", "B", "C") if c in abc)
    corpo = f"""<div class="caixa-aviso"><b>Minuta para revisão.</b> Consolidação
automática do que o município já contratou, segundo os registros do PNCP.
Os itens publicados <b>não trazem código de catálogo</b> (CATMAT/CATSER),
exigido no plano oficial — a classificação, o agrupamento definitivo e a
conferência de especificação e unidade cabem ao gestor.<br>
Quantidade pela <b>{base}</b>, acrescida da margem informada; preço unitário
pela <b>{est}</b> dos valores homologados.</div>
<div class="cards">
<div class="card"><div class="n">{d['totais']['grupos']}</div><div class="l">itens no plano</div></div>
<div class="card"><div class="n">{moeda(d['totais']['valor'])}</div><div class="l">valor estimado</div></div>
<div class="card"><div class="n">{p.get('margem', '—')}%</div><div class="l">margem aplicada</div></div>
</div>{f'<p class="disp">Curva ABC — {linha_abc}. Classe A concentra 80% do valor, B os 15% seguintes: é onde a revisão rende mais.</p>' if linha_abc else ''}
<h2>Itens da minuta</h2>
<table><thead><tr><th class="num">#</th><th>Descrição</th>
<th class="ctr">Tipo</th><th class="ctr">Unid.</th><th class="num">Quantidade</th>
<th class="num">Unitário</th><th class="num">Total</th>
<th class="ctr" title="Curva ABC: A concentra 80% do valor, B os 15% seguintes, C o resto">ABC</th></tr></thead>
<tbody>{linhas or '<tr><td colspan="8">Minuta vazia.</td></tr>'}</tbody></table>"""
    titulo = f"{TITULOS['minuta_pca']} {d['ano']} — {municipio}"
    return _pagina(titulo, corpo, municipio, uf, f"Exercício {d['ano']}",
                   paisagem=True, brasao=brasao, categoria=categoria,
                   acervo=acervo)


def _desconsiderados_html(d):
    """Seção dos preços que ficaram de fora, com a razão de cada um.

    O documento tem de bastar a si mesmo: quem confere precisa ver que a
    série foi filtrada, o que saiu e por quê. Item sem razão registrada
    aparece marcado — é pendência antes de assinar, não detalhe.
    """
    fora = d.get("desconsiderados") or []
    if not fora:
        return ""
    sem_motivo = sum(1 for l in fora if not l.get("motivo"))
    linhas = "".join(f"""<tr>
      <td class="obj">{_e(l['descricao'])}</td>
      <td class="unid">{_e(l['unidade'])}</td>
      <td class="num">{quantidade(l['quantidade_homologada'])}</td>
      <td class="num">{moeda(l['valor_unitario_homologado'])}</td>
      <td class="forn">{_e(l['fornecedor_nome'])}</td>
      <td class="ctr proc">{_processo(l)}</td>
      <td class="{'sem-motivo' if not l.get('motivo') else ''}">{
        _e(l['motivo']) if l.get('motivo')
        else 'Sem justificativa registrada'}</td></tr>"""
      for l in fora)
    alerta = (f" <b>{sem_motivo} {'item' if sem_motivo == 1 else 'itens'} "
              f"{'está' if sem_motivo == 1 else 'estão'} sem justificativa "
              "registrada</b> — registre a razão antes de juntar este "
              "documento ao processo." if sem_motivo else "")
    return f"""<h2>Itens desconsiderados nesta pesquisa</h2>
<div class="caixa-aviso">Os preços abaixo foram coletados, mas <b>não entraram
no cálculo</b> por decisão do responsável pela pesquisa. Eles constam aqui para
que a filtragem fique visível a quem confere.{alerta}</div>
<table><thead><tr><th>Descrição</th><th class="unid">Unid.</th>
<th class="num">Qtde</th><th class="num">Unitário</th>
<th class="forn">Fornecedor</th><th class="ctr">Processo</th>
<th>Razão</th></tr></thead><tbody>{linhas}</tbody></table>"""


def _processo(l):
    """Número do processo com link para o PNCP, quando dá para montar."""
    texto = f"{_e(l['sequencial'])}/{_e(l['ano'])}"
    url = url_pncp(l.get("orgao_cnpj"), l.get("ano"), l.get("sequencial"))
    return (f'<a href="{_e(url)}" title="Abrir no PNCP">{texto}</a>'
            if url else texto)






def render_executivo(d, municipio, uf, brasao=None, graficos=None,
                     categoria=None, acervo=None):
    """Reformulado (2026-08-08, pedido do usuário) para usar os mesmos
    gráficos do Painel — hero com sparkline, colunas mensais pareadas e
    barras por modalidade — em vez de só tabelas. `d` é o retorno de
    `dados_painel`: mesma consulta, mesmos números do que está na tela.

    `graficos` (achado 2026-08-11, mesmo padrão da pesquisa de preços):
    dict `{"meses": html, "modalidade": html}` com o SVG que a tela já
    desenhou em ECharts. Slot ausente ou `graficos=None` cai no
    `_grafico_meses`/`_grafico_barras` de sempre — chamada direta, CLI e
    testes continuam funcionando sem depender de navegador nenhum.
    """
    graficos = graficos or {}
    ano = d["ano"]
    ex = d["execucao"]
    c = ex["cards"]
    desagio = f"{c['desagio']:.1f}%".replace(".", ",") \
        if c["desagio"] is not None else "–"
    ate_hoje = " até hoje" if d["comparacao_parcial"] else ""

    var_valor = None
    if c["homologado"] and ex["homologado_anterior"]:
        var_valor = (c["homologado"] / ex["homologado_anterior"] - 1) * 100
    if var_valor is None:
        linha_valor = f"sem {ano - 1} para comparar"
    else:
        seta = "▲" if var_valor >= 0 else "▼"
        classe = "up" if var_valor >= 0 else "down"
        pct_txt = f"{abs(var_valor):.1f}%".replace(".", ",")
        linha_valor = (f'<span class="{classe}">{seta} {pct_txt}</span>'
                       f' sobre {ano - 1}{ate_hoje}')
    var_n = c["n"] - (ex["n_anterior"] or 0)
    seta_n = "▲" if var_n >= 0 else "▼"

    economia = ""
    if c.get("estimado") and c.get("homologado"):
        economia = f"{moeda(c['estimado'] - c['homologado'])} economizados"

    # sparkline do hero: mesmo traçado de ui/painel.js:vistaExecucao
    pontos_spark = [m["valor"] for m in ex["meses"] if m["valor"]]
    spark = ""
    if len(pontos_spark) > 1:
        maxs = max(pontos_spark) or 1
        n = len(pontos_spark)
        linha_pts = ",".join(
            f"{8 + i * (224 / max(1, n - 1)):.1f},{38 - (v / maxs) * 32:.1f}"
            for i, v in enumerate(pontos_spark))
        spark = _svg(240, 44, f'<polyline fill="none" stroke="var(--s1)"'
                              f' stroke-width="2" stroke-linejoin="round"'
                              f' points="{linha_pts}"/>')

    hero = f"""<div class="faixa f-4">
<div class="card hero">
  <h3>Homologado em {ano}</h3>
  <div class="n">{moeda(c['homologado'])}</div>
  <div class="r">{linha_valor}</div>
  {spark}
</div>
<div class="card kpiv"><div class="v">{c['n']}</div>
  <div class="r">contratações</div>
  <div class="r" style="margin-top:8px">{seta_n} {abs(var_n)} vs.
    {ano - 1}{ate_hoje}</div></div>
<div class="card kpiv"><div class="v">{desagio}</div>
  <div class="r">deságio médio</div>
  <div class="r" style="margin-top:8px">{economia}</div></div>
<div class="card kpiv"><div class="v">{c['contratos_vigentes']}</div>
  <div class="r">contratos vigentes</div>
  <div class="r" style="margin-top:8px">{c['atas_vigentes']} atas vigentes</div>
</div>
</div>"""

    graf_meses = graficos.get("meses") or _grafico_meses(
        ex["meses"], "var(--s1)", larg=580)
    graf_mod = graficos.get("modalidade") or _grafico_barras(
        ex["modalidades"][:6],
        valor=lambda m: m["homologado"] or m["estimado"] or 0,
        rotulo=lambda m: m["modalidade_nome"] or "–",
        sub=lambda m: f"{m['n']} {'processo' if m['n'] == 1 else 'processos'}",
        cor="var(--s1)", larg=340)
    charts = f"""<div class="faixa f-21">
<div class="card"><h3>Contratações por mês — estimado × homologado</h3>
{graf_meses}</div>
<div class="card"><h3>Por modalidade — valor homologado</h3>
{graf_mod}</div>
</div>"""

    mod = "".join(f"""<tr><td>{_e(m['modalidade_nome'])}</td>
      <td class="num">{m['n']}</td>
      <td class="num">{moeda(m['estimado'])}</td>
      <td class="num">{moeda(m['homologado'])}</td></tr>"""
      for m in ex["modalidades"])
    meses_por_n = {m["mes"]: m for m in ex["meses"]}
    meses = "".join(f"""<tr><td class="ctr">{MESES_NOME[i-1]}</td>
      <td class="num">{meses_por_n.get(i, {}).get('n', 0)}</td>
      <td class="num">{moeda(meses_por_n[i]['valor']) if meses_por_n.get(i, {}).get('valor') else '–'}</td></tr>"""
      for i in range(1, 13))
    forn = "".join(f"""<tr><td>{_e(f['fornecedor_nome'])}<br>
      <small>{_e(documento(f['fornecedor_ni']))}</small></td>
      <td class="num">{f['n']}</td><td class="num">{moeda(f['total'])}</td></tr>"""
      for f in ex["fornecedores"])
    venc = "".join(f"""<tr><td class="ctr">{_e(v['tipo'])}</td>
      <td class="ctr">{_e(v['nome'])}</td>
      <td class="obj">{_e(v['objeto'])}</td>
      <td class="num">{data_br(v['vigencia_fim'])}</td>
      <td class="num">{v['dias']} dias</td></tr>"""
      for v in ex["vencendo"])
    corpo = f"""<div class="caixa-aviso">Resumo executivo da execução do
exercício, extraído do PNCP — contratações publicadas, deságio sobre o
estimado e prazos vencendo. A comparação com o exercício anterior é
cortada no mesmo dia do calendário, para não medir ano inteiro contra ano
em curso.</div>
{hero}
{charts}
<h2>Contratações por modalidade</h2>
<table><thead><tr><th>Modalidade</th><th class="num">Qtde</th>
<th class="num">Estimado</th><th class="num">Homologado</th></tr></thead>
<tbody>{mod or '<tr><td colspan="4">Sem dados.</td></tr>'}</tbody></table>
<h2>Evolução mensal (valor homologado/estimado publicado)</h2>
<table><thead><tr><th class="ctr">Mês</th><th class="num">Processos</th>
<th class="num">Valor</th></tr></thead><tbody>{meses}</tbody></table>
<h2>Maiores fornecedores contratados no ano</h2>
<table><thead><tr><th>Fornecedor</th><th class="num">Contratos</th>
<th class="num">Valor</th></tr></thead>
<tbody>{forn or '<tr><td colspan="3">Sem contratos no ano.</td></tr>'}</tbody></table>
<h2>Vigências a vencer nos próximos 90 dias</h2>
<table><thead><tr><th class="ctr">Tipo</th><th class="ctr">Contrato/Ata</th><th>Objeto</th>
<th class="num">Fim</th><th class="num">Prazo</th></tr></thead>
<tbody>{venc or '<tr><td colspan="5">Nada vence nos próximos 90 dias.</td></tr>'}</tbody></table>"""
    titulo = f"{TITULOS['executivo']} {ano} — {municipio}"
    return _pagina(titulo, corpo, municipio, uf, f"Exercício {ano}",
                   paisagem=True, estilo_extra=CSS_PAINEL,
                   brasao=brasao, categoria=categoria, acervo=acervo)


def render_economia(d, municipio, uf, brasao=None, graficos=None,
                    categoria=None, acervo=None):
    """Quanto foi economizado no ano, por modalidade, família de item e
    categoria do PNCP. `d` é o retorno de `dados_painel` — mesmos números
    da vista Economia do Painel, num documento que se gera sem abrir o
    Painel.

    `graficos`: mesmo dicionário-opcional que `render_executivo` aceita —
    `{"modalidade": html, "familia": html, "categoria": html,
    "fornecedor": html}`. Slot ausente cai no `_grafico_barras` de sempre.
    """
    graficos = graficos or {}
    ano = d["ano"]
    e = d["economia"]
    ate_hoje = " até hoje" if d["comparacao_parcial"] else ""

    var_econ = None
    if e.get("economizado_anterior"):
        var_econ = (e["economizado"] / e["economizado_anterior"] - 1) * 100
    if var_econ is None:
        linha_econ = f"sem {ano - 1} para comparar"
    else:
        seta = "▲" if var_econ >= 0 else "▼"
        classe = "up" if var_econ >= 0 else "down"
        pct_txt = f"{abs(var_econ):.1f}%".replace(".", ",")
        linha_econ = (f'<span class="{classe}">{seta} {pct_txt}</span>'
                      f' sobre {ano - 1}{ate_hoje}')
    desagio = f"{e['pct']:.1f}%".replace(".", ",") if e["pct"] is not None \
        else "–"

    hero = f"""<div class="faixa f-3">
<div class="card hero">
  <h3>Economizado em {ano}</h3>
  <div class="n">{moeda(e['economizado'])}</div>
  <div class="r">{linha_econ}</div>
</div>
<div class="card kpiv"><div class="v">{desagio}</div>
  <div class="r">deságio médio</div>
  <div class="r" style="margin-top:8px">{moeda(e['estimado'])} estimados</div>
</div>
<div class="card kpiv"><div class="v">{moeda(e['homologado'])}</div>
  <div class="r">homologado no ano</div></div>
</div>"""

    graf_mod = graficos.get("modalidade") or _grafico_barras(
        e["por_modalidade"], valor=lambda m: m["economizado"] or 0,
        rotulo=lambda m: m["modalidade"] or "–",
        sub=lambda m: f"{m['n']} {'processo' if m['n'] == 1 else 'processos'}",
        cor="var(--s1)", larg=300)
    graf_fam = graficos.get("familia") or _grafico_barras(
        e["por_familia"], valor=lambda f: f["economizado"] or 0,
        rotulo=lambda f: f["nome"] or "–",
        sub=lambda f: f"{f['n']} {'item' if f['n'] == 1 else 'itens'}",
        cor="var(--s1)", larg=300)
    graf_cat = graficos.get("categoria") or _grafico_barras(
        e["por_categoria"], valor=lambda c: c["economizado"] or 0,
        rotulo=lambda c: c["nome"] or "–",
        sub=lambda c: f"{c['n']} {'item' if c['n'] == 1 else 'itens'}",
        cor="var(--s1)", larg=300)
    graf_forn = graficos.get("fornecedor") or _grafico_barras(
        e["por_fornecedor"], valor=lambda f: f["economizado"] or 0,
        rotulo=lambda f: f["nome"] or "–",
        sub=lambda f: f"{f['n']} {'item' if f['n'] == 1 else 'itens'} · "
                      f"{f['pct']:.0f}%",
        cor="var(--s1)", larg=900)
    charts = f"""<div class="faixa f-3">
<div class="card"><h3>Por modalidade</h3>{graf_mod}</div>
<div class="card"><h3>Por família de item</h3>{graf_fam}</div>
<div class="card"><h3>Por categoria (PNCP)</h3>{graf_cat}</div>
</div>
<div class="card"><h3>Por fornecedor — quem fechou abaixo do estimado</h3>
{graf_forn}
<div class="nota">Agrupado pelo CNPJ/CPF, não pelo nome — a mesma empresa
aparece com grafias diferentes entre processos. Deságio alto não é atestado
de bom fornecedor: pode ser estimativa inflada na origem.</div></div>"""

    def _tabela(titulo_secao, coluna, linhas, rotulo=None):
        corpo_linhas = "".join(f"""<tr>
          <td>{rotulo(l) if rotulo else _e(l['nome'])}</td>
          <td class="num">{l['n']}</td>
          <td class="num">{moeda(l['estimado'])}</td>
          <td class="num">{moeda(l['homologado'])}</td>
          <td class="num">{moeda(l['economizado'])}</td></tr>"""
          for l in linhas)
        return f"""<h2>{titulo_secao}</h2>
<table><thead><tr><th>{coluna}</th><th class="num">Qtde</th>
<th class="num">Estimado</th><th class="num">Homologado</th>
<th class="num">Economizado</th></tr></thead>
<tbody>{corpo_linhas or f'<tr><td colspan="5">Sem dados.</td></tr>'}</tbody>
</table>"""

    mod_linhas = [{"nome": m["modalidade"], **m} for m in e["por_modalidade"]]
    corpo = f"""<div class="caixa-aviso">Economia por modalidade, família de
item, categoria e fornecedor, extraída do PNCP — mede o que foi
<b>estimado</b> contra o que foi de fato <b>homologado</b>. Não é auditoria
de preço: estimativa alta nem sempre significa preço de mercado mal
calculado, e homologado sem estimativa comparável fica de fora da
soma.</div>
{hero}
{charts}
{_tabela("Economia por modalidade", "Modalidade", mod_linhas)}
{_tabela("Economia por família de item", "Família", e["por_familia"])}
{_tabela("Economia por categoria (PNCP)", "Categoria", e["por_categoria"])}
{_tabela("Economia por fornecedor", "Fornecedor", e["por_fornecedor"],
         rotulo=lambda f: f"{_e(f['nome'])}<br>"
                          f"<small>{_e(documento(f['ni']))}</small>")}"""
    titulo = f"{TITULOS['economia']} {ano} — {municipio}"
    return _pagina(titulo, corpo, municipio, uf, f"Exercício {ano}",
                   paisagem=True, estilo_extra=CSS_PAINEL,
                   brasao=brasao, categoria=categoria, acervo=acervo)


# ── geração (HTML + CSV) ────────────────────────────────────────────────────

# Cores de série do papel. São as do tema Portal de ui/estilo.css — o único
# conjunto calibrado para superfície branca, que é a do documento desde a
# v1.20.0. Os conjuntos do Pergaminho e do Observatório continuam existindo
# na tela (ui/estilo.css); aqui não entram porque, medidos contra papel
# branco, caem a 1,28-2,99 de contraste. Detalhe em design/DASHBOARD.md.
SERIE_DOCUMENTO = dict(s1="#2a78d6", s2="#eb6834", s3="#1baf7a", s4="#eda100",
                       seq1="#cde2fb", seq2="#9ec5f4", seq3="#5598e7",
                       seq4="#2571c9", seq5="#1c5cab")

# Tinta do número dentro da célula do mapa de calor. Precisa vir para o papel
# junto da rampa: sem ela o número herda a cor de texto do documento e some
# nos dois degraus escuros. Chave com hífen não cabe em `dict(...)`, daí o
# literal. Os cinco pares foram medidos contra a rampa acima (≥ 4,5:1) — o
# `seq4` mudou de #2a78d6 para #2571c9 justamente porque, no anterior, nem o
# branco puro alcançava 4,5:1.
TINTA_SEQUENCIAL = {"seq1-ink": "#141414", "seq2-ink": "#141414",
                    "seq3-ink": "#141414", "seq4-ink": "#ffffff",
                    "seq5-ink": "#ffffff"}


# Estilo do painel impresso. As cores de série são as mesmas da tela — foram
# validadas para daltonismo e contraste —, e `print-color-adjust: exact` é o
# que impede o navegador de "economizar tinta" e devolver barras cinzentas.
_CSS_PAINEL_RESTO = """
  * { print-color-adjust: exact; -webkit-print-color-adjust: exact; }
  /* O SVG capturado chega com viewBox (posto na captura, ver paraPapel em
     ui/painel.js) e largura de 100%: estas duas linhas são o que faz a
     altura acompanhar, em vez de o desenho ser cortado no cartão estreito. */
  .card svg { max-width:100%; height:auto; display:block; }
  /* o par gráfico + overlay do corte vertical dividem a mesma caixa; sem
     isto o overlay fica na altura antiga e as marcas descolam do gráfico */
  .card svg[data-overlay] { position:absolute; inset:0; }
  /* A tela grava a altura do gráfico em pixels no `style=` do contêiner,
     medida para a largura da tela. No papel o SVG passa a ocupar 100% da
     largura e fica MAIS ALTO que aquele número — vazava a caixa e caía por
     cima da nota do cartão. `auto` deixa a caixa acompanhar o desenho; o
     `!important` é para vencer o style inline, que não dá para remover sem
     mexer no que a tela usa. */
  .card .graf, .card .graf-echart { height:auto !important; }
  /* Gráfico com overlay (o corte vertical desenhado à mão) põe os rótulos
     de eixo ABAIXO da caixa do gráfico. Na tela a altura fixa do contêiner
     já reservava esse espaço; com `height:auto` a caixa encolheu até o
     desenho e os rótulos passaram a encostar na nota do cartão — medido em
     -5 px de folga. O respiro devolve o que a altura fixa dava. */
  .card .graf-par { padding-bottom:14px; }
  .vista { display:grid; gap:12px; }
  .faixa { display:grid; gap:12px; }
  .f-4 { grid-template-columns:1.15fr 1fr 1fr 1fr; }
  .f-21 { grid-template-columns:1.6fr 1fr; }
  .f-11 { grid-template-columns:1fr 1fr; }
  .f-3 { grid-template-columns:1fr 1fr 1fr; }
  /* min-width:0 devolve ao grid o direito de apertar o item abaixo do
     min-content (a largura fixa do SVG do ECharts). Sem isso, no papel a
     faixa transborda a página e os gráficos da direita saem cortados —
     espelha estilo.css:561 (a tela ganhou isso na 1.42.0, o papel não). */
  .faixa > *, .card, .graf, .graf-par, .graf-echart { min-width:0; }
  .card { background:var(--superficie); border:1px solid var(--borda);
          border-top:3px solid var(--cor-categoria, var(--acento));
          border-radius:3px; padding:11px 14px 12px; break-inside:avoid; }
  .card h3 { font-size:9.5pt; color:var(--suave); font-weight:600;
             letter-spacing:.05em; text-transform:uppercase; margin-bottom:8px; }
  .hero .n { font-size:26pt; font-weight:700; line-height:1.05; }
  .hero .r, .kpiv .r { font-size:9pt; color:var(--suave); margin-top:2px; }
  .kpiv .v { font-size:16pt; font-weight:700; }
  .kpiv .r { text-transform:uppercase; letter-spacing:.05em; font-size:8pt; }
  .up { color:#2f7d32; font-weight:600; } .down { color:var(--alerta); font-weight:600; }
  .leg { display:flex; gap:14px; font-size:8.5pt; color:var(--suave);
         margin-top:6px; }
  .leg i { width:9px; height:9px; border-radius:2px; display:inline-block;
           margin-right:5px; }
  .nota { font-size:8.5pt; color:var(--suave); margin-top:7px; line-height:1.45; }
  .vazio { color:var(--suave); font-size:9pt; padding:18px 0; text-align:center; }
  .rot { font-size:8pt; fill:var(--suave); }
  .val { font-size:8.5pt; fill:var(--texto); }
  .eixo { stroke:var(--borda); stroke-width:1; }
  .badge { font-size:8pt; padding:2px 8px; border-radius:99px; }
  .badge.ok { background:#e6f4ea; color:#2f7d32; }
  .badge.warn { background:#fdf1dc; color:var(--atencao); }
  .badge.err { background:#fbe9e7; color:var(--alerta); }
  /* Calendário da agenda. PRECISA estar aqui: o painel impresso não carrega
     o ui/estilo.css — ele leva só o HTML das vistas e é este bloco que o
     formata. Sem estas regras a grade some e os 92 dias saem empilhados
     numa coluna, ocupando duas páginas (achado ao conferir o PDF real,
     2026-08-14; eu havia escrito no DESIGN.md que "sai no papel sem
     conversão" sem imprimir para conferir). */
  .cal { display:grid; grid-template-columns:repeat(3, 1fr); gap:14px; }
  .cal-mes h4 { font-size:8pt; letter-spacing:.08em; text-transform:uppercase;
                color:var(--suave); font-weight:600; margin-bottom:5px; }
  .cal-sem, .cal-grade { display:grid; grid-template-columns:repeat(7, 1fr);
                         gap:2px; }
  .cal-sem { margin-bottom:2px; }
  .cal-sem span { font-size:7pt; color:var(--suave); text-align:center; }
  .cal-dia { aspect-ratio:1; border-radius:3px; background:var(--cabecalho);
             display:grid; place-items:center; font-size:7.5pt;
             color:var(--suave); position:relative; }
  .cal-dia.fora { background:none; }
  .cal-dia.venc { font-weight:700; }
  .cal-dia.venc.u { background:#fbe9e7; color:var(--alerta); }
  .cal-dia.venc.a { background:#fdf1dc; color:var(--atencao); }
  .cal-dia.venc.t { background:#e6f4ea; color:#2f7d32; }
  .cal-dia b { position:absolute; top:-4px; right:-4px; min-width:13px;
               height:13px; padding:0 3px; border-radius:99px; font-size:7pt;
               line-height:13px; text-align:center; font-weight:700;
               color:#fff; border:1px solid var(--superficie); }
  .cal-dia.venc.u b { background:var(--alerta); }
  .cal-dia.venc.a b { background:var(--atencao); }
  .cal-dia.venc.t b { background:#2f7d32; }
  /* o contorno de "hoje" é orientação de tela; no papel a data já está lá */
  .cal-dia.hoje { outline:none; }
  /* instrução de clique não faz sentido no papel */
  .so-tela { display:none; }
  .secao-painel { break-after:page; }
  /* :last-of-type, não :last-child — depois da última seção vem o <footer>,
     então a última seção NÃO é o último filho da página. Com :last-child o
     seletor não casava, a última vista mantinha o break-after:page e empurrava
     o rodapé sozinho para uma página em branco no fim (PDF real, 2026-08-16). */
  .secao-painel:last-of-type { break-after:auto; }
  .secao-painel > h2 { font-family:Georgia,serif; font-size:15pt;
                       font-weight:400; color:var(--acento); margin:0 0 10px; }
  .chips { display:flex; gap:8px; flex-wrap:wrap; margin-bottom:12px; }
  .chip { font-size:9pt; padding:5px 12px; border-radius:99px;
          border:1px solid var(--borda); background:var(--superficie); }
"""


# As chaves de SERIE_DOCUMENTO já são os nomes das variáveis CSS (s1…seq5).
CSS_PAINEL = (
    ":root { "
    + " ".join(f"--{chave}:{cor};" for chave, cor
               in {**SERIE_DOCUMENTO, **TINTA_SEQUENCIAL}.items())
    + " --surface:var(--superficie); --surface2:var(--cabecalho);"
    " --muted:var(--suave); --text:var(--texto); --border:var(--borda);"
    " --accent:var(--acento); --accent-fg:#ffffff; --erro:var(--alerta);"
    " --warn:var(--atencao); --ok:#2f7d32; --pill:99px;"
    " --font-ui:'Segoe UI',system-ui,sans-serif; }") + _CSS_PAINEL_RESTO


TITULOS_PAINEL = {"execucao": "Execução do exercício",
                  "analise": "Análise comparativa",
                  "vigilancia": "Vigilância e prazos",
                  "economia": "Economia e comparativos"}


def render_painel(vistas, municipio, uf, ano, brasao=None):
    """Monta o painel impresso a partir do que a tela desenhou.

    Os gráficos não são redesenhados aqui: o SVG que vai ao papel é o mesmo
    que está na tela, enviado pela interface. Redesenhar no Python seria uma
    segunda implementação para divergir da primeira.
    """
    corpo = "".join(
        f'<section class="secao-painel"><h2>{_e(TITULOS_PAINEL.get(nome, nome))}'
        f'</h2>{html}</section>'
        for nome, html in vistas if html)
    return _pagina(f"Painel — {municipio} — {ano}", corpo, municipio, uf,
                   # A4 deitado: é o papel que toda impressora de
                   # secretaria tem. Quem quiser A3 pede na caixa de
                   # impressão do navegador, e o desenho acompanha —
                   # a grade é fluida, não tem medida travada em pixel.
                   f"Exercício {ano}", paisagem=True, papel="A4",
                   estilo_extra=CSS_PAINEL, brasao=brasao)


CSS_FICHA = """
  .ficha-objeto { text-transform:uppercase; text-align:justify;
                  hyphens:auto; font-family:Georgia,serif; font-size:14.5px;
                  line-height:1.5; color:var(--texto); margin:2px 0 20px; }
  .ficha-grid { display:grid; grid-template-columns:repeat(2,1fr);
                gap:10px 24px; margin-top:4px; }
  .ficha-grid > div { break-inside:avoid; }
  .ficha-grid .k { font-size:9.5px; letter-spacing:.05em;
                   text-transform:uppercase; color:var(--suave); }
  .ficha-grid .v { font-size:12.5px; color:var(--texto); margin-top:1px; }
  .ficha-grid .v a { color:var(--acento); text-decoration:none;
                      border-bottom:1px dotted var(--acento); }
  .ficha-raw { margin-top:22px; padding-top:14px;
               border-top:1px solid var(--borda); }
  .ficha-raw h2 { margin-bottom:8px; }
  .ficha-raw pre { background:var(--cabecalho); border:1px solid var(--borda);
                   border-radius:3px; padding:10px 12px; font-size:10px;
                   line-height:1.5; white-space:pre-wrap; word-break:break-word; }
  .j-chave { color:var(--acento); }
  .j-str   { color:#2f7d32; }
  .j-num   { color:var(--atencao); }
  .j-bool, .j-null { color:var(--suave); font-weight:600; }
"""


def render_detalhe(titulo, subtitulo, meta_html, municipio, uf, brasao=None,
                    raw_html="", titulo_doc=None):
    """Ficha impressa de um registro específico — contratação, contrato,
    ata, item ou PCA — aberto no modal de detalhe da tela.

    Cabeçalho é o brasão + identificação do município (mesmo padrão dos
    demais relatórios) — o objeto não cabia ali sem afogar o município,
    então desce pro corpo como parágrafo em caixa alta e justificado
    (pedido do usuário, achado 2026-08-12), com a grade de campos abaixo.

    `meta_html`/`raw_html` vêm prontos do que a tela já mostra (rótulo/
    valor formatados e o JSON colorido do "Dados completos"): mesmo
    padrão dos demais relatórios (tela desenha, papel captura), sem
    reimplementar rótulo por rótulo nem o realce do JSON aqui.

    `titulo_doc`: nome do documento (vira `<title>` — o nome sugerido
    ao "Salvar como PDF"). Sem ele, cai no padrão "Município — UF".
    """
    corpo = f'<p class="ficha-objeto">{_e(titulo)}</p>'
    corpo += f'<div class="ficha-grid">{meta_html}</div>'
    if raw_html:
        corpo += (f'<div class="ficha-raw"><h2>Dados completos (JSON do '
                  f'PNCP)</h2><pre>{raw_html}</pre></div>')
    return _pagina(titulo_doc or f"{municipio} — {uf}", corpo, municipio, uf,
                   subtitulo or "", paisagem=False, estilo_extra=CSS_FICHA,
                   brasao=brasao)


def gerar(db, tipo, params, municipio, uf, destino):
    """Gera o relatório e retorna {"html": caminho, "csv": caminho|None}."""
    params = params or {}
    ano = params.get("ano")
    vigentes = bool(params.get("vigentes"))
    orgao = params.get("orgao")
    # com filtro de órgão, o nome dele entra no cabeçalho, no título
    # (= nome do PDF) e no nome do arquivo
    if orgao and params.get("orgao_nome"):
        municipio = f"{municipio} · {params['orgao_nome']}"
    destino.mkdir(parents=True, exist_ok=True)
    # brasão do município (Configurações) — toma o lugar do estandarte no
    # cabeçalho de todo documento gerado aqui, se tiver sido enviado
    linha_brasao = db.execute(
        "SELECT valor FROM config WHERE chave='brasao'").fetchone()
    brasao = linha_brasao[0] if linha_brasao else None
    # selo de procedência (achado 2026-08-13): categoria e acervo passam
    # por todo `render_*` até `_pagina`, igual `brasao` — mesma fotografia
    # do acervo em toda a tarja/faixa/rodapé de um mesmo documento
    categoria = CATEGORIA_RELATORIO.get(tipo)
    acervo = _acervo_atual(db)
    if tipo == "executivo":
        if not ano:
            ano = date.today().year
        d = dados_painel(db, ano, orgao, params.get("limites"))
        conteudo = render_executivo(d, municipio, uf, brasao=brasao,
                                    graficos=params.get("graficos"),
                                    categoria=categoria, acervo=acervo)
        nome = f"resumo_executivo_{ano}"
        linhas_csv = None
    elif tipo == "economia":
        if not ano:
            ano = date.today().year
        d = dados_painel(db, ano, orgao, params.get("limites"))
        conteudo = render_economia(d, municipio, uf, brasao=brasao,
                                   graficos=params.get("graficos"),
                                   categoria=categoria, acervo=acervo)
        nome = f"economia_comparativo_{ano}"
        linhas_csv = d["economia"]["por_familia"]
    elif tipo == "minuta_pca":
        if not ano:
            ano = date.today().year + 1
        itens = pca_builder.listar_minuta(db, ano, so_incluidos=True)
        cfg = db.execute("SELECT parametros FROM pca_minuta WHERE ano_alvo=?",
                         (ano,)).fetchone()
        d = {"ano": ano, "itens": itens,
             "totais": pca_builder.totais(itens),
             "parametros": json.loads(cfg[0]) if cfg else {}}
        conteudo = render_minuta_pca(d, municipio, uf, brasao=brasao,
                                     categoria=categoria, acervo=acervo)
        nome = f"minuta_pca_{ano}"
        linhas_csv = [{k: i[k] for k in ("descricao", "unidade", "categoria",
                                         "quantidade", "valor_unitario",
                                         "margem", "valor_total")}
                      for i in itens]
    elif tipo == "fracionamento":
        if not ano:
            ano = date.today().year
        d = dados_fracionamento(db, ano, orgao, params.get("limites"))
        conteudo = render_fracionamento(d, municipio, uf, brasao=brasao,
                                        categoria=categoria, acervo=acervo)
        nome = f"alerta_fracionamento_{ano}"
        linhas_csv = d["dispensas"]
    else:
        periodo_txt = ("Vigentes em " + date.today().strftime("%d/%m/%Y")) \
            if vigentes else (f"Exercício {ano}" if ano else "Todo o período")
        if tipo == "contratacoes":
            d = dados_contratacoes(db, ano, params.get("modalidade"), orgao)
            conteudo = render_contratacoes(d, municipio, uf, periodo_txt,
                                           brasao=brasao, categoria=categoria,
                                           acervo=acervo)
        elif tipo == "contratos":
            d = dados_contratos(db, ano, vigentes, orgao)
            conteudo = render_contratos(d, municipio, uf, periodo_txt,
                                        brasao=brasao, categoria=categoria,
                                        acervo=acervo)
        elif tipo == "atas":
            d = dados_atas(db, ano, vigentes, orgao)
            conteudo = render_atas(d, municipio, uf, periodo_txt,
                                   brasao=brasao, categoria=categoria,
                                   acervo=acervo)
        else:
            raise ValueError(f"tipo de relatório desconhecido: {tipo}")
        sufixo = "vigentes" if vigentes else (str(ano) if ano else "completo")
        nome = f"relacao_{tipo}_{sufixo}"
        linhas_csv = d["linhas"]
    if orgao:
        nome += f"_orgao_{orgao}"
    caminho_html = destino / f"{nome}.html"
    caminho_html.write_text(conteudo, encoding="utf-8")
    caminho_xlsx = None
    if linhas_csv:
        caminho_xlsx = destino / f"{nome}.xlsx"
        escrever_planilha(caminho_xlsx, linhas_csv)
    return {"html": str(caminho_html),
            "xlsx": str(caminho_xlsx) if caminho_xlsx else None}
